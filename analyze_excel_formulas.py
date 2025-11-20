"""
Excel Formula Analysis Script
Analyzes the Intprepix Volume III.xlsx file to understand:
1. Wrap rate calculation sequence (DL → Fringe → OH → G&A → ODC → Fee)
2. Actual formulas in cells
3. Sheet 2 rate breakdown calculations
4. Escalation implementation
5. Formula vs hardcoded values
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
from collections import defaultdict

def analyze_cell(cell):
    """Extract information about a cell including its formula and value."""
    info = {
        'coordinate': cell.coordinate,
        'value': cell.value,
        'data_type': cell.data_type,
        'has_formula': False,
        'formula': None,
        'number_format': cell.number_format
    }

    # Check if cell has a formula (data_type 'f' means formula)
    if cell.data_type == 'f':
        info['has_formula'] = True
        info['formula'] = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None

    return info

def find_wrap_rate_patterns(sheet):
    """Find patterns in wrap rate calculations."""
    patterns = {
        'direct_labor': [],
        'fringe': [],
        'overhead': [],
        'g_and_a': [],
        'odc': [],
        'fee': [],
        'escalation': []
    }

    # Search for common keywords in formulas
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formula = str(cell.value).upper()

                # Check for various calculation components
                if 'FRINGE' in formula or any(x in formula for x in ['*0.', '*(1+', '*1.']):
                    patterns['fringe'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

                if 'OVERHEAD' in formula or 'OH' in formula:
                    patterns['overhead'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

                if 'G&A' in formula or 'GA' in formula or 'G AND A' in formula:
                    patterns['g_and_a'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

                if 'ODC' in formula:
                    patterns['odc'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

                if 'FEE' in formula:
                    patterns['fee'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

                if 'ESCALAT' in formula or '1.02' in formula or '1.03' in formula:
                    patterns['escalation'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

    return patterns

def analyze_sheet(sheet, sheet_name):
    """Comprehensive analysis of a sheet."""
    print(f"\n{'='*80}")
    print(f"ANALYZING SHEET: {sheet_name}")
    print(f"{'='*80}\n")

    # Get dimensions
    print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns\n")

    # Collect all formulas
    formula_cells = []
    value_cells = []

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_info = analyze_cell(cell)
                if cell_info['has_formula']:
                    formula_cells.append(cell_info)
                else:
                    value_cells.append(cell_info)

    print(f"Total cells with formulas: {len(formula_cells)}")
    print(f"Total cells with hardcoded values: {len(value_cells)}\n")

    # Show header row (typically row 1)
    print("HEADER ROW (Row 1):")
    print("-" * 80)
    for cell in sheet[1]:
        if cell.value:
            print(f"  {cell.coordinate}: {cell.value}")

    # Show first few rows of data with formulas
    print("\n\nFIRST 10 ROWS WITH FORMULAS:")
    print("-" * 80)
    shown = 0
    for formula_info in formula_cells[:50]:  # Check first 50 formula cells
        if shown < 10:
            print(f"\nCell {formula_info['coordinate']}:")
            print(f"  Formula: {formula_info['formula']}")
            print(f"  Result: {formula_info['value']}")
            print(f"  Format: {formula_info['number_format']}")
            shown += 1

    # Find wrap rate patterns
    print("\n\nWRAP RATE CALCULATION PATTERNS:")
    print("-" * 80)
    patterns = find_wrap_rate_patterns(sheet)

    for component, matches in patterns.items():
        if matches:
            print(f"\n{component.upper().replace('_', ' ')} ({len(matches)} occurrences):")
            for match in matches[:3]:  # Show first 3 examples
                print(f"  {match['cell']}: {match['formula']}")

    # Look for rate columns
    print("\n\nSEARCHING FOR RATE/PERCENTAGE COLUMNS:")
    print("-" * 80)
    for col_idx in range(1, min(sheet.max_column + 1, 30)):
        col_letter = get_column_letter(col_idx)
        header_cell = sheet[f"{col_letter}1"]
        if header_cell.value:
            header_text = str(header_cell.value).upper()
            if any(keyword in header_text for keyword in ['RATE', 'PERCENT', '%', 'FRINGE', 'OH', 'G&A', 'ODC', 'FEE', 'ESCALAT']):
                print(f"\nColumn {col_letter}: {header_cell.value}")
                # Show a few examples from this column
                for row_idx in range(2, min(7, sheet.max_row + 1)):
                    cell = sheet[f"{col_letter}{row_idx}"]
                    if cell.value:
                        cell_info = analyze_cell(cell)
                        if cell_info['has_formula']:
                            print(f"  {cell.coordinate}: Formula = {cell_info['formula']}")
                        else:
                            print(f"  {cell.coordinate}: Value = {cell.value}")

    return {
        'formula_cells': formula_cells,
        'value_cells': value_cells,
        'patterns': patterns
    }

def trace_calculation_sequence(sheet):
    """Trace the sequence of wrap rate calculations."""
    print("\n\n" + "="*80)
    print("TRACING WRAP RATE CALCULATION SEQUENCE")
    print("="*80)

    # Look for a complete calculation chain in one row
    print("\nLooking for calculation chains in rows...")

    for row_idx in range(2, min(20, sheet.max_row + 1)):
        row = sheet[row_idx]
        row_formulas = []

        for cell in row:
            if cell.data_type == 'f' and cell.value:
                row_formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'value': cell.value
                })

        if len(row_formulas) >= 3:  # If row has multiple formulas
            print(f"\nRow {row_idx} calculation chain:")
            for formula_info in row_formulas[:10]:
                print(f"  {formula_info['cell']}: {formula_info['formula']}")

def main():
    file_path = '/Users/keshav/Developer/Others/Pricing/Intprepix Volume III.xlsx'

    print("="*80)
    print("EXCEL FORMULA ANALYZER")
    print("="*80)
    print(f"\nFile: {file_path}\n")

    try:
        # Load workbook with data_only=False to get formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)

        print(f"Sheet names: {wb.sheetnames}\n")

        # Analyze each sheet
        results = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            results[sheet_name] = analyze_sheet(sheet, sheet_name)

            # Trace calculation sequence
            trace_calculation_sequence(sheet)

        # Summary Report
        print("\n\n" + "="*80)
        print("SUMMARY REPORT")
        print("="*80)

        for sheet_name, data in results.items():
            print(f"\n{sheet_name}:")
            print(f"  Formula cells: {len(data['formula_cells'])}")
            print(f"  Value cells: {len(data['value_cells'])}")
            print(f"  Pattern matches:")
            for pattern_name, matches in data['patterns'].items():
                if matches:
                    print(f"    {pattern_name}: {len(matches)}")

        # Key findings
        print("\n\nKEY FINDINGS:")
        print("-" * 80)
        print("\nTo determine ODC placement in wrap rate sequence:")
        print("1. Look for formulas that reference previous calculations")
        print("2. Identify the order: DL → DL+Fringe → +OH → +G&A → +ODC → +Fee")
        print("3. Check if ODC is applied before or after G&A")
        print("4. Verify if Fee is calculated on all costs including ODC")

        print("\n\nFormula patterns to look for:")
        print("- Sequential multiplication: base * (1 + rate1) * (1 + rate2)")
        print("- Cumulative addition: base + fringe + OH + G&A + ODC")
        print("- Cell references showing calculation chain")

    except Exception as e:
        print(f"Error analyzing file: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
