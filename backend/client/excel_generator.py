"""
Excel Generator for Government Contract Cost Proposals.
Refactored to match the Price IQ Sample Template format.

Creates professional Excel files with:
- Multiple sheets (CE Summary, Labor Detail, Subcontractor sheets, Material, Travel, LOE, Indirect Rates)
- Blue header styling (#284C82)
- Thin borders on all cells
- Period-based columns with support for Base, Options, and Extensions
"""

from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .calculation_service import Calculator


def format_soc_code(soc_code: str) -> str:
    """
    Format SOC code to consistent 6-digit display format without hyphen.

    Examples:
    - "15-1252" → "151252"
    - "151252" → "151252" (already formatted)
    - None → "-"

    Args:
        soc_code: Raw SOC code (6 digits with or without hyphen)

    Returns:
        Formatted SOC code without hyphen (XXXXXX) or "-" if invalid
    """
    if not soc_code:
        return '-'

    # Remove any existing hyphens
    clean = soc_code.replace('-', '')

    # Validate: must be 6 digits
    if len(clean) != 6 or not clean.isdigit():
        return soc_code  # Return as-is if invalid format

    # Return as 6-digit format without hyphen
    return clean


class ExcelGenerator:
    """
    Generates government contract cost proposal Excel files.
    Matches the Price IQ Sample Template format.
    """

    # Style constants matching sample template
    HEADER_FILL = PatternFill(start_color="284C82", end_color="284C82", fill_type="solid")  # Dark blue
    PERIOD_HEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # Light blue
    HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Exact formats from template
    CURRENCY_FORMAT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
    PERCENT_FORMAT = '0.00%'
    NUMBER_FORMAT = '#,##0.00'  # Template uses decimal for hours

    # Indirect Rate sheet — fixed row positions (column B holds the value)
    # These are referenced by ALL other sheets via _ir_ref()
    IR_FRINGE_ROW = 9
    IR_OH_ONSITE_ROW = 10
    IR_OH_OFFSITE_ROW = 11
    IR_GA_ROW = 12
    IR_PASSTHROUGH_ROW = 13         # S&MH + G&A Passthrough — for subcontractor passthrough line only
    IR_FEE_LABOR_ROW = 14
    IR_FEE_SUB_ROW = 15
    IR_SMH_MATERIALS_ROW = 16       # S&MH only — for materials/ODC handling (matches Nexagen sample)
    # Escalation rows start at 17 (row 16 is blank)
    IR_ESCALATION_START_ROW = 17

    # Helper: compute prime overtime $ per year (OT hours × fee-inclusive FBLR × OT multiplier)
    def _compute_prime_ot_by_year(self):
        """Compute prime overtime cost per year by iterating positions.

        Mirrors PRICING_FORMULAS.md § 16.1: per-position OT cost per year =
            ot_hours × FBLR(year) × ot_multiplier
        where FBLR is fee-inclusive and includes compound escalation.

        Returns {1: ot_cost_year_1, 2: …, …} — literal dollar amounts that the
        caller writes directly into Excel cells (no cross-sheet formula here
        because the PLD sheet lacks per-position OT columns).
        """
        from client.calculation_service import Calculator  # local import to avoid cycles

        indirect_rates = self.project_data.get('indirect_rates', {}) or {}
        escalation_rates = self.project_data.get('escalation_rates', {}) or {}
        fee_rates = self.project_data.get('fee_rates', {}) or {}

        ot_multiplier = indirect_rates.get('ot_multiplier') or 1.5
        fringe_rate = indirect_rates.get('fringe', 0) or 0
        oh_onsite = indirect_rates.get('oh_onsite', indirect_rates.get('oh', 0)) or 0
        oh_offsite = indirect_rates.get('oh_offsite', indirect_rates.get('oh', 0)) or 0
        ga_rate = indirect_rates.get('ga', 0) or 0
        fee_rate = fee_rates.get('prime_labor', 0) or 0

        result = {y: 0.0 for y in range(1, self.total_years + 1)}
        for pos in self.project_data.get('prime_positions', []):
            ot_map = pos.get('ot_hours_per_year') or {}
            if not ot_map:
                continue
            is_gsa = (pos.get('wage_source') or '').lower() == 'gsa'

            if is_gsa:
                gsa_rates = pos.get('gsa_rates_by_year') or {}
                gsa_current_year = pos.get('gsa_current_year') or 1
                discount = pos.get('gsa_discount_rate') or 0.0
                for y in range(1, self.total_years + 1):
                    ot_hours = float(ot_map.get(str(y)) or 0)
                    if ot_hours <= 0:
                        continue
                    rate = Calculator.get_gsa_rate_for_year(
                        gsa_rates_by_year=gsa_rates,
                        gsa_current_year=gsa_current_year,
                        proposal_year=y,
                        escalation_rates=escalation_rates,
                        gsa_custom_rate=pos.get('gsa_custom_rate'),
                    )
                    rate *= (1 - discount)
                    result[y] += ot_hours * rate * ot_multiplier
            else:
                base_wage = pos.get('base_annual_wage', 0) or 0
                fte = pos.get('standard_fte_hours') or 1920
                loc = pos.get('location_type', 'On-Site')
                oh_rate = oh_onsite if loc == 'On-Site' else oh_offsite
                for y in range(1, self.total_years + 1):
                    ot_hours = float(ot_map.get(str(y)) or 0)
                    if ot_hours <= 0:
                        continue
                    # Compound-escalate wage to target year
                    wage = base_wage
                    for yy in range(1, y):
                        esc = escalation_rates.get(f"{yy}_to_{yy + 1}", 0) or 0
                        wage *= (1 + esc)
                    dl_rate = wage / fte
                    fringe_amt = dl_rate * fringe_rate
                    oh_amt = (dl_rate + fringe_amt) * oh_rate
                    ga_amt = (dl_rate + fringe_amt + oh_amt) * ga_rate
                    fee_amt = (dl_rate + fringe_amt + oh_amt + ga_amt) * fee_rate
                    fblr = dl_rate + fringe_amt + oh_amt + ga_amt + fee_amt
                    result[y] += ot_hours * fblr * ot_multiplier

        return result

    def __init__(self):
        """Initialize the Excel generator."""
        self.wb = None
        self.total_years = 0
        self.project_data = None
        self.extensions = []

    def _ir_ref(self, row: int) -> str:
        """Return an absolute cell reference into the Indirect Rate sheet (column B)."""
        return f"'Indirect Rate'!$B${row}"

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize a string for use as an Excel sheet name (max 31 chars, no special chars)."""
        return name[:31].replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')

    def _escalation_formula(self, from_year: int, to_year: int) -> str:
        """
        Build a compound escalation multiplier fragment referencing Indirect Rate sheet.
        E.g. from_year=1, to_year=3 → '*(1+IR!$B$17)*(1+IR!$B$18)'
        Returns empty string when from_year >= to_year (year 1, no escalation needed).
        """
        parts = []
        for y in range(from_year, to_year):
            ir_row = self.IR_ESCALATION_START_ROW + (y - 1)
            parts.append(f"*(1+{self._ir_ref(ir_row)})")
        return "".join(parts)

    def generate_cost_proposal(self, project_data: Dict[str, Any]) -> Workbook:
        """
        Generate complete cost proposal Excel workbook in Price IQ format.

        Sheet creation order is designed for true feeder architecture:
          - CE Summary worksheet is created first (tab 0) but filled last
          - Indirect Rate sheet is built next (feeder 1: all rate %)
          - Prime Labor Detail is built next (year 1 rates = BLS/GSA values,
            year N rates = Excel formulas referencing IR escalation rows)
          - FLLR is built after PLD (feeder 2: BLS FBLR = PLD rate × IR cascade)
          - ODC / Materials / Travel built next (base+handling split, escalation via IR)
          - CE Summary is filled last, referencing all other sheets via formulas

        Args:
            project_data: Complete project data including all positions, rates, etc.

        Returns:
            Workbook ready to save
        """
        self.wb = Workbook()
        self.project_data = project_data
        self.total_years = project_data['total_years']
        self.extensions = project_data.get('extensions', [])

        # True when any prime position uses a GSA catalog rate
        self.has_gsa = any(
            (p.get('wage_source') or '').lower() == 'gsa'
            for p in project_data.get('prime_positions', [])
        )

        # Surge option — percentage (e.g. 0.20) and multiplier (e.g. 1.15)
        surge = project_data.get('surge')
        self.surge_percentage = (surge.get('percentage') or 0) if surge else 0
        self.surge_multiplier = project_data.get('surge_multiplier') or 1.15
        self.ir_surge_multiplier_row = None  # set by _create_indirect_rates_sheet
        # Analysis sheet name: "GSA Analysis" for GSA proposals, "BLS Analysis" for BLS
        self.analysis_sheet_name = "GSA Analysis" if self.has_gsa else "BLS Analysis"
        # PLD year-col start: col 5 (E) when GSA Discount col D exists, else col 4 (D)
        self.pld_year_start = 5 if self.has_gsa else 4
        # Standard FTE hours — same for all positions, read once from the first position
        _first = next(iter(project_data.get('prime_positions', [])), {})
        self.standard_fte_hours = _first.get('standard_fte_hours') or 1920

        # 1. CE Summary — create worksheet first so it stays as tab 0; fill content last
        ws_summary = self.wb.active
        ws_summary.title = "Summary"
        self._setup_ce_summary_columns(ws_summary)

        # 2. Indirect Rate — feeder sheet 1 (all rate % values in fixed rows)
        self._create_indirect_rates_sheet()

        # 3. Prime Labor Detail — DL rates (year 1 = BLS/GSA value, year N = IR escalation formula)
        pld_info = self._create_prime_labor_detail_sheet()

        # 4. Subcontractor sheets — one per subcontractor, named by company
        sub_infos = []
        if project_data.get('subcontractors'):
            for sub in project_data['subcontractors']:
                sub_info = self._create_subcontractor_sheet(sub)
                sub_infos.append(sub_info)

        # 5. Fully Loaded Labor Rates — feeder sheet 2
        #    BLS: PLD year N rate × IR burden cascade (PLD yr1 feeds from BLS Analysis)
        #    GSA: direct reference to PLD (already fully loaded)
        #    Sub: direct reference to sub sheet rate column
        self._create_fully_loaded_labor_rates_sheet()

        # 6. ODCs — base + handling rows; base year N = IR escalation formula
        odc_info = None
        if project_data.get('odcs'):
            odc_info = self._create_odcs_sheet()

        # 7. Materials — base + handling rows; base year N = IR escalation formula
        materials_info = None
        if project_data.get('materials'):
            materials_info = self._create_materials_sheet()

        # 8. Travel — base + G&A rows; base year N = IR escalation formula
        travel_info = None
        if project_data.get('travel'):
            travel_info = self._create_travel_sheet()

        # 9. LOE sheet
        self._create_loe_sheet()

        # 10. BLS Analysis sheet
        if project_data.get('wage_data'):
            self._create_bls_analysis_sheet()

        # 11. Fill CE Summary last — every value references another sheet via formula
        self._fill_ce_summary_sheet(ws_summary, pld_info, sub_infos, odc_info, materials_info, travel_info)

        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        return self.wb

    def _apply_standard_header(self, ws, start_col=1):
        """
        Apply standard header format to a worksheet.

        Creates table-style header with labels and values in separate cells:
        Row 1: "Proprietary Data" (red, bold, centered, merged)
        Row 2: Prime Contractor Name | [value]
        Row 3: Subcontractor(s) Name | [value]
        Row 4: Solicitation | [value]
        Row 5: Task Order Number | [value]

        Args:
            ws: Worksheet to apply header to
            start_col: Starting column (default 1 = A)
        """
        # Row 1: "Proprietary Data" - Red, Bold, Centered
        proprietary_cell = ws.cell(1, start_col, "Proprietary Data")
        proprietary_cell.font = Font(bold=True, size=11, color="FF0000")
        proprietary_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Get subcontractor names
        sub_names = ", ".join(self.project_data.get('subcontractor_names', []))
        if not sub_names:
            sub_names = ""

        # Table-style header rows (2-5): Label in col A, Value in col B onwards
        header_data = [
            ("Prime Contractor Name", self.project_data.get('prime_contractor_name', 'N/A')),
            ("Subcontractor(s) Name", sub_names),
            ("Solicitation", self.project_data.get('solicitation_number', 'N/A')),
            ("Task Order Number", self.project_data.get('task_order_number', ''))
        ]

        for idx, (label, value) in enumerate(header_data):
            row = 2 + idx

            # Label column (bold, bordered)
            label_cell = ws.cell(row, start_col, label)
            label_cell.font = self.BOLD_FONT
            label_cell.border = self.THIN_BORDER
            label_cell.alignment = Alignment(horizontal='left', vertical='center')

            # Value column (bordered, next column)
            value_cell = ws.cell(row, start_col + 1, value)
            value_cell.border = self.THIN_BORDER
            value_cell.alignment = Alignment(horizontal='left', vertical='center')

    def _get_period_label(self, year: int) -> str:
        """Get label for a given year (Base, Option, Extension)."""
        base_years = self.project_data.get('base_years', 1)

        if year <= base_years:
            if base_years == 1:
                return "Base Period"
            return f"Base Year {year}"

        # Check if it's an extension
        for ext in self.extensions:
            if ext['year'] == year:
                return ext.get('label', f"Extension {year}")

        # Regular option year
        option_num = year - base_years
        return f"Option Period {option_num}"

    def _setup_ce_summary_columns(self, ws):
        """Set up CE Summary worksheet column widths and standard header (rows 1-5).
        Called early so the sheet exists as tab 0; data is filled later by _fill_ce_summary_sheet."""

        # Set column widths matching template exactly
        # A: Cost Element Labels
        ws.column_dimensions['A'].width = 40.66
        # B: First period column
        ws.column_dimensions['B'].width = 20.66
        # Remaining period columns
        for col_idx in range(3, 2 + self.total_years + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
        # Total column
        ws.column_dimensions[get_column_letter(2 + self.total_years)].width = 13

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

    def _fill_ce_summary_sheet(self, ws, pld_info, sub_infos, odc_info, materials_info, travel_info):
        """
        Fill CE Summary with cross-sheet formula references.

        Non-GSA (BLS) layout — 13 data rows:
          Direct Labor | Fringe Benefits | Labor Overhead | G&A (Labor)
          Subcontractor(s) | Passthrough (S&MH + G&A)
          Materials | Materials Handling
          Travel | G&A (Travel)
          Sub-Total | Fee | Total Proposed

        GSA layout — simplified single Prime Labor row (legacy behavior).

        All values are Excel formulas referencing BLS Analysis / PLD / sub sheets / IR sheet.
        """
        header_row = 9
        data_start_row = 10
        current_row = data_start_row

        # Total col = one column after the last period
        total_col = 2 + self.total_years

        # ── Row 9: Column headers ─────────────────────────────────────────────
        ws.cell(header_row, 1, "Cost Element")
        self._style_header_cell(ws.cell(header_row, 1))
        col = 2
        for year in range(1, self.total_years + 1):
            period_cell = ws.cell(header_row, col, self._get_period_label(year))
            self._style_period_header_cell(period_cell)
            col += 1
        ws.cell(header_row, col, "Total")
        self._style_header_cell(ws.cell(header_row, col))

        # ── Helper: write one CE Summary row ──────────────────────────────────
        def _write_row(label, year_formulas, total_formula, bold=True):
            """Write a labelled row with per-year formulas and a total formula."""
            lbl_cell = ws.cell(current_row, 1, label)
            lbl_cell.border = self.THIN_BORDER
            if bold:
                lbl_cell.font = self.BOLD_FONT
            for period_idx, formula in enumerate(year_formulas):
                c = ws.cell(current_row, 2 + period_idx, formula)
                c.number_format = self.CURRENCY_FORMAT
                c.border = self.THIN_BORDER
                if bold:
                    c.font = self.BOLD_FONT
            tc = ws.cell(current_row, total_col, total_formula)
            tc.number_format = self.CURRENCY_FORMAT
            tc.border = self.THIN_BORDER
            tc.font = self.BOLD_FONT

        # ── PLD dollar-column helper (year N → col letter in PLD sheet) ───────
        def _pld_col(year):
            return get_column_letter(3 * year + self.pld_year_start - 1)

        pld_tcl = get_column_letter(pld_info['total_dollars_col']) if pld_info else 'B'

        if self.has_gsa:
            # ── GSA: single "Prime Labor" row referencing PLD total ───────────
            # Sub positions are rendered directly in PLD (for GSA only), so the
            # PLD total naturally includes all labor costs — no separate sub row.
            if pld_info and pld_info.get('total_prime_labor_row'):
                r = pld_info['total_prime_labor_row']
                year_fmls = [f"='Prime Labor Detail'!{_pld_col(y)}{r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Prime Labor Detail'!{pld_tcl}{r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Prime Labor", year_fmls, total_fml)
            prime_labor_row_num = current_row
            current_row += 1

            # Surge Option (GSA)
            if self.surge_percentage:
                mult_ref = self._ir_ref(self.ir_surge_multiplier_row)
                surge_year_fmls = [
                    f"={get_column_letter(2 + i)}{prime_labor_row_num}*{self.surge_percentage}*{mult_ref}"
                    for i in range(self.total_years)
                ]
                surge_total_fml = f"={get_column_letter(total_col)}{prime_labor_row_num}*{self.surge_percentage}*{mult_ref}"
                _write_row(
                    f"Surge Option ({self.surge_percentage * 100:.1f}% × {(self.surge_multiplier - 1) * 100:.1f}% premium)",
                    surge_year_fmls, surge_total_fml, bold=False
                )
                current_row += 1

        else:
            # ── BLS: full DCAA cost-element breakdown ─────────────────────────

            # 1. Direct Labor — track CE row so Fee can reference it
            dl_row_num = current_row
            if pld_info and pld_info.get('total_dl_row'):
                r = pld_info['total_dl_row']
                year_fmls = [f"='Prime Labor Detail'!{_pld_col(y)}{r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Prime Labor Detail'!{pld_tcl}{r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Direct Labor", year_fmls, total_fml)
            current_row += 1

            # 2. Fringe Benefits
            if pld_info and pld_info.get('fringe_row'):
                r = pld_info['fringe_row']
                year_fmls = [f"='Prime Labor Detail'!{_pld_col(y)}{r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Prime Labor Detail'!{pld_tcl}{r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Fringe Benefits", year_fmls, total_fml, bold=False)
            current_row += 1

            # 3. Labor Overhead
            if pld_info and pld_info.get('oh_row'):
                r = pld_info['oh_row']
                year_fmls = [f"='Prime Labor Detail'!{_pld_col(y)}{r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Prime Labor Detail'!{pld_tcl}{r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Labor Overhead", year_fmls, total_fml, bold=False)
            current_row += 1

            # 4. General & Administrative (Labor) — track CE row (last prime-labor row)
            ga_ce_row_num = current_row
            if pld_info and pld_info.get('ga_row'):
                r = pld_info['ga_row']
                year_fmls = [f"='Prime Labor Detail'!{_pld_col(y)}{r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Prime Labor Detail'!{pld_tcl}{r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("General & Administrative (Labor)", year_fmls, total_fml, bold=False)
            current_row += 1

            # 4b. Overtime (Prime) — placed after G&A (Labor) but BEFORE sub row
            # so it's swept into the Sub-Total sum. NOT in the fee base, because
            # OT is already fee-inclusive (FBLR × OT hours × multiplier).
            ot_ce_row_num = current_row
            if pld_info and pld_info.get('overtime_row'):
                r = pld_info['overtime_row']
                year_fmls = [f"='Prime Labor Detail'!{_pld_col(y)}{r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Prime Labor Detail'!{pld_tcl}{r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Overtime (Prime)", year_fmls, total_fml, bold=False)
            current_row += 1

            # Surge Option (BLS) — based on fully burdened prime labor (DL:G&A × (1 + fee_rate))
            # This matches the UI: surge is on FBLR (includes fee), not just bare labor burden
            if self.surge_percentage:
                mult_ref = self._ir_ref(self.ir_surge_multiplier_row)
                fee_ref = self._ir_ref(self.IR_FEE_LABOR_ROW)
                surge_year_fmls = []
                for i in range(self.total_years):
                    cl = get_column_letter(2 + i)
                    surge_year_fmls.append(
                        f"=SUM({cl}{dl_row_num}:{cl}{ga_ce_row_num})*(1+{fee_ref})*{self.surge_percentage}*{mult_ref}"
                    )
                surge_total_fml = (
                    f"=SUM({get_column_letter(total_col)}{dl_row_num}:{get_column_letter(total_col)}{ga_ce_row_num})"
                    f"*(1+{fee_ref})*{self.surge_percentage}*{mult_ref}"
                )
                _write_row(
                    f"Surge Option ({self.surge_percentage * 100:.1f}% × {(self.surge_multiplier - 1) * 100:.1f}% premium)",
                    surge_year_fmls, surge_total_fml, bold=False
                )
                current_row += 1

        # ── Subcontractor(s): sum sub sheet total-rows ────────────────────────
        # GSA: sub positions already appear in Prime Labor (shows_in_main_grid),
        # so the sub tab is internal-only and we exclude this row from CE Summary.
        sub_row_num = current_row
        if not self.has_gsa:
            if sub_infos:
                year_fmls = []
                for year in range(1, self.total_years + 1):
                    sub_dollars_col = get_column_letter(3 * year + 2)
                    parts = [f"'{si['sheet_name']}'!{sub_dollars_col}{si['total_row']}" for si in sub_infos]
                    year_fmls.append("=" + "+".join(parts))
                total_parts = [f"'{si['sheet_name']}'!{get_column_letter(si['total_col'])}{si['total_row']}" for si in sub_infos]
                total_fml = "=" + "+".join(total_parts)
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Subcontractor(s)", year_fmls, total_fml)
            current_row += 1

        # ── Passthrough (S&MH + G&A): Sub row × IR!B13 ───────────────────────
        if not self.has_gsa:
            pass_year_fmls = [
                f"={get_column_letter(2 + i)}{sub_row_num}*{self._ir_ref(self.IR_PASSTHROUGH_ROW)}"
                for i in range(self.total_years)
            ]
            pass_total_fml = f"={get_column_letter(total_col)}{sub_row_num}*{self._ir_ref(self.IR_PASSTHROUGH_ROW)}"
            _write_row("Passthrough (S&MH + G&A)", pass_year_fmls, pass_total_fml, bold=False)
            current_row += 1

        # ── Materials: ODC base-only total (or 0) ────────────────────────────
        if not self.has_gsa:
            if odc_info and odc_info.get('base_only_total_row'):
                odc_base_r = odc_info['base_only_total_row']
                year_fmls = [f"='ODCs'!{get_column_letter(y + 1)}{odc_base_r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='ODCs'!{get_column_letter(odc_info['total_col'])}{odc_base_r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Materials", year_fmls, total_fml, bold=False)
            current_row += 1

            # ── Materials Handling: ODC handling row (or 0) ───────────────────
            if odc_info and odc_info.get('handling_row'):
                odc_h_r = odc_info['handling_row']
                year_fmls = [f"='ODCs'!{get_column_letter(y + 1)}{odc_h_r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='ODCs'!{get_column_letter(odc_info['total_col'])}{odc_h_r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
            _write_row("Materials Handling", year_fmls, total_fml, bold=False)
            current_row += 1

        # ── Travel ────────────────────────────────────────────────────────────
        # GSA: use total_travel_row (Base + G&A combined) — no separate G&A row below
        # BLS: use base_total_row only — G&A on Travel gets its own row below
        travel_row_num = current_row
        if travel_info:
            tr_ref_r = (
                travel_info.get('total_travel_row')
                if self.has_gsa
                else travel_info.get('base_total_row')
            )
            if tr_ref_r:
                year_fmls = [f"='Travel'!{get_column_letter(y + 1)}{tr_ref_r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Travel'!{get_column_letter(travel_info['total_col'])}{tr_ref_r}"
            else:
                year_fmls = [0] * self.total_years
                total_fml = 0
        else:
            year_fmls = [0] * self.total_years
            total_fml = 0
        _write_row("Travel", year_fmls, total_fml)
        current_row += 1

        # ── G&A (Travel): Travel sheet ga_on_travel_row (or inline formula) ──
        if not self.has_gsa:
            if travel_info and travel_info.get('ga_on_travel_row'):
                ga_tr_r = travel_info['ga_on_travel_row']
                year_fmls = [f"='Travel'!{get_column_letter(y + 1)}{ga_tr_r}" for y in range(1, self.total_years + 1)]
                total_fml = f"='Travel'!{get_column_letter(travel_info['total_col'])}{ga_tr_r}"
            else:
                # Fallback: compute from travel row in this sheet × G&A rate
                year_fmls = [
                    f"={get_column_letter(2 + i)}{travel_row_num}*{self._ir_ref(self.IR_GA_ROW)}"
                    for i in range(self.total_years)
                ]
                total_fml = f"={get_column_letter(total_col)}{travel_row_num}*{self._ir_ref(self.IR_GA_ROW)}"
            _write_row("General & Administrative (Travel)", year_fmls, total_fml, bold=False)
            current_row += 1

        # ── Sub-Total: SUM of all data rows ──────────────────────────────────
        sub_total_row_num = current_row
        for period_idx in range(self.total_years):
            cl = get_column_letter(2 + period_idx)
            c = ws.cell(current_row, 2 + period_idx)
            c.value = f"=SUM({cl}{data_start_row}:{cl}{current_row - 1})"
            c.number_format = self.CURRENCY_FORMAT
            c.border = self.THIN_BORDER
            c.font = self.BOLD_FONT
        lbl = ws.cell(current_row, 1, "Sub-Total")
        lbl.font = self.BOLD_FONT
        lbl.border = self.THIN_BORDER
        tcl = get_column_letter(total_col)
        tc = ws.cell(current_row, total_col)
        tc.value = f"=SUM({tcl}{data_start_row}:{tcl}{current_row - 1})"
        tc.number_format = self.CURRENCY_FORMAT
        tc.border = self.THIN_BORDER
        tc.font = self.BOLD_FONT
        current_row += 1

        # ── GSA disclaimer ───────────────────────────────────────────────────
        if self.has_gsa and sub_infos:
            current_row += 1  # blank row after Sub-Total
            sub_sheet_names = ", ".join(f'"{si["sheet_name"]}"' for si in sub_infos)
            disclaimer_text = (
                f"NOTE: The {sub_sheet_names} sheet(s) are for internal reference only. "
                f"It is recommended to delete them before submitting this proposal."
            )
            cell = ws.cell(current_row, 1, disclaimer_text)
            cell.font = Font(italic=True, color="FF0000", size=10)
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=total_col
            )

        # ── Fee (non-GSA only) ────────────────────────────────────────────────
        # Formula mirrors the UI exactly:
        #   Fee = SUM(DL:G&A) × fee_on_labor_rate  +  Sub × fee_on_sub_rate
        # Travel, passthrough, materials are NOT in the fee base — matching
        # calculateGrandTotal() in pricingStore.ts (primeFee + subFee).
        if not self.has_gsa:
            fee_row_num = current_row
            fee_year_fmls = []
            for i in range(self.total_years):
                cl = get_column_letter(2 + i)
                prime_labor_sum = f"SUM({cl}{dl_row_num}:{cl}{ga_ce_row_num})"
                sub_ref = f"{cl}{sub_row_num}"
                fee_year_fmls.append(
                    f"={prime_labor_sum}*{self._ir_ref(self.IR_FEE_LABOR_ROW)}"
                    f"+{sub_ref}*{self._ir_ref(self.IR_FEE_SUB_ROW)}"
                )
            tcl = get_column_letter(total_col)
            prime_labor_sum_total = f"SUM({tcl}{dl_row_num}:{tcl}{ga_ce_row_num})"
            fee_total_fml = (
                f"={prime_labor_sum_total}*{self._ir_ref(self.IR_FEE_LABOR_ROW)}"
                f"+{tcl}{sub_row_num}*{self._ir_ref(self.IR_FEE_SUB_ROW)}"
            )
            _write_row("Fee", fee_year_fmls, fee_total_fml, bold=False)
            current_row += 1

            # ── Total Proposed: Sub-Total + Fee ──────────────────────────────
            tp_year_fmls = [
                f"={get_column_letter(2 + i)}{sub_total_row_num}+{get_column_letter(2 + i)}{fee_row_num}"
                for i in range(self.total_years)
            ]
            tp_total_fml = f"={get_column_letter(total_col)}{sub_total_row_num}+{get_column_letter(total_col)}{fee_row_num}"
            _write_row("Total Proposed", tp_year_fmls, tp_total_fml)

    def _create_prime_labor_detail_sheet(self):
        """Create the Prime Labor Detail sheet for prime contractor positions with FBLR breakdown."""
        ws = self.wb.create_sheet("Prime Labor Detail")

        # Column widths matching template
        # A: Labor Category (was B)
        ws.column_dimensions['A'].width = 50.5
        # B: Site (was C)
        ws.column_dimensions['B'].width = 19.66
        # C: Location (was D)
        ws.column_dimensions['C'].width = 14.33

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Column headers row
        header_row = 10
        headers = ["Labor Category", "Site", "Location"]
        if self.has_gsa:
            headers.append("GSA Discount")
            ws.column_dimensions['D'].width = 14
        for idx, header in enumerate(headers):
            cell = ws.cell(header_row, 1 + idx)
            cell.value = header
            self._style_header_cell(cell)

        # For each period: Hours, Rate, Dollars — start at pld_year_start
        col = self.pld_year_start
        for year in range(1, self.total_years + 1):
            period_label = self._get_period_label(year)

            # Period header above sub-headers (row 7) - Light blue background, merged across 3 columns
            period_cell = ws.cell(7, col, period_label)
            self._style_period_header_cell(period_cell)
            ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 2)

            # Sub-headers for Hours, Rate, Dollars
            sub_headers = ["Hours/Base", "Rate", "Dollars"]
            for sub_idx, sub_header in enumerate(sub_headers):
                cell = ws.cell(header_row, col + sub_idx)
                cell.value = sub_header
                self._style_header_cell(cell)

                # Set column widths matching template
                if sub_idx == 0:  # Hours
                    ws.column_dimensions[get_column_letter(col + sub_idx)].width = 14.33
                elif sub_idx == 1:  # Rate
                    ws.column_dimensions[get_column_letter(col + sub_idx)].width = 11.33
                else:  # Dollars
                    ws.column_dimensions[get_column_letter(col + sub_idx)].width = 14.5

            col += 3

        # Total columns
        ws.cell(header_row, col, "Total Hours")
        self._style_header_cell(ws.cell(header_row, col))
        total_hours_col = col
        col += 1

        ws.cell(header_row, col, "Total Dollars")
        self._style_header_cell(ws.cell(header_row, col))
        total_dollars_col = col

        # Data rows
        current_row = header_row + 1
        positions = self.project_data.get('prime_positions', [])

        for position in positions:
            ws.cell(current_row, 1, position['labor_category'])
            ws.cell(current_row, 1).border = self.THIN_BORDER

            ws.cell(current_row, 2, position.get('site', 'Government'))
            ws.cell(current_row, 2).border = self.THIN_BORDER

            ws.cell(current_row, 3, position.get('location', ''))
            ws.cell(current_row, 3).border = self.THIN_BORDER

            # Col D: GSA Discount — only written when proposal has GSA positions
            if self.has_gsa:
                discount_cell = ws.cell(current_row, 4)
                if (position.get('wage_source') or '').lower() == 'gsa':
                    discount_cell.value = position.get('gsa_discount_rate', 0.0) or 0.0
                    discount_cell.number_format = self.PERCENT_FORMAT
                discount_cell.border = self.THIN_BORDER

            # Calculate position data for each year
            # Use appropriate calculator based on wage source
            if (position.get('wage_source') or '').lower() == 'gsa':
                # GSA positions: Use GSA-specific calculator
                discount_rate = position.get('gsa_discount_rate', 0.0)
                results = Calculator.calculate_gsa_position_years(
                    position_data=position,
                    total_years=self.total_years,
                    discount_rate=discount_rate,
                    escalation_rates=self.project_data.get('escalation_rates', {})
                )
            else:
                # BLS positions: Use standard calculator
                results = Calculator.calculate_position_years(
                    position_data=position,
                    escalation_rates=self.project_data['escalation_rates'],
                    indirect_rates=self.project_data['indirect_rates'],
                    total_years=self.total_years
                )

            col = self.pld_year_start
            hours_cells = []
            dollars_cells = []

            for year in range(1, self.total_years + 1):
                year_data = results.get(f'year_{year}', {})

                # Hours
                hours_col = col
                cell = ws.cell(current_row, col)
                cell.value = year_data.get('hours', 0)
                cell.number_format = self.NUMBER_FORMAT
                cell.border = self.THIN_BORDER
                hours_cells.append(f"{get_column_letter(col)}{current_row}")
                col += 1

                # Rate column
                # GSA yr N: BLS Analysis col L(yr1)/M(yr2)/N(yr3)... × (1-discount) / burden_factor
                #   BLS Analysis stores actual GSA catalog rates per year; for years beyond
                #   available data it escalates via IR sheet. Dividing by burden_factor gives
                #   the DL component so the FBLR cascade restores the full GSA hourly total.
                # BLS yr 1: direct BLS Analysis row ref ÷ hours (no duplicate-name bug)
                # BLS yr N: escalate from year 1 rate cell via IR escalation formulas
                rate_col = col
                cell = ws.cell(current_row, col)
                # year1 rate col = pld_year_start + 1
                year1_rate_col = get_column_letter(self.pld_year_start + 1)
                if (position.get('wage_source') or '').lower() == 'gsa':
                    bls_row = position.get('bls_analysis_row')
                    if bls_row:
                        # GSA: store fully-loaded rate × (1-discount) directly — no burden division
                        bls_col = get_column_letter(11 + year - 1)
                        cell.value = (
                            f"='{self.analysis_sheet_name}'!${bls_col}${bls_row}*(1-$D{current_row})"
                        )
                    else:
                        cell.value = year_data.get('rate', 0)
                elif year == 1:
                    if self.project_data.get('wage_data'):
                        bls_row = position.get('bls_analysis_row')
                        bls_wage_col = 'K' if self.has_gsa else 'L'
                        if bls_row:
                            cell.value = f"='{self.analysis_sheet_name}'!${bls_wage_col}${bls_row}/{self.standard_fte_hours}"
                        else:
                            cell.value = (
                                f"=INDEX('{self.analysis_sheet_name}'!${bls_wage_col}:${bls_wage_col},"
                                f"MATCH(A{current_row},'{self.analysis_sheet_name}'!$A:$A,0))"
                                f"/{self.standard_fte_hours}"
                            )
                    else:
                        cell.value = year_data.get('dl_rate', 0)
                else:
                    cell.value = f"={year1_rate_col}{current_row}{self._escalation_formula(1, year)}"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                col += 1

                # Dollars - USE FORMULA: Hours * Rate
                cell = ws.cell(current_row, col)
                cell.value = f"={get_column_letter(hours_col)}{current_row}*{get_column_letter(rate_col)}{current_row}"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                dollars_cells.append(f"{get_column_letter(col)}{current_row}")
                col += 1

            # Total Hours
            cell = ws.cell(current_row, total_hours_col)
            cell.value = f"={'+'.join(hours_cells)}"
            cell.number_format = self.NUMBER_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT

            # Total Dollars
            cell = ws.cell(current_row, total_dollars_col)
            cell.value = f"={'+'.join(dollars_cells)}"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT

            current_row += 1

        # ── GSA only: render sub positions directly in PLD ────────────────────
        # Sub positions are internal-only; showing them here folds their costs
        # into the PLD total so CE Summary "Prime Labor" captures everything.
        if self.has_gsa:
            for sub in self.project_data.get('subcontractors', []):
                for labor_cat in sub.get('labor_categories', []):
                    ws.cell(current_row, 1, labor_cat['labor_category'])
                    ws.cell(current_row, 1).border = self.THIN_BORDER
                    ws.cell(current_row, 2, labor_cat.get('site', 'Government'))
                    ws.cell(current_row, 2).border = self.THIN_BORDER
                    ws.cell(current_row, 3, labor_cat.get('location', ''))
                    ws.cell(current_row, 3).border = self.THIN_BORDER

                    # Col D: discount — sub rates are already discounted, show 0
                    discount_cell = ws.cell(current_row, 4)
                    discount_cell.value = 0.0
                    discount_cell.number_format = self.PERCENT_FORMAT
                    discount_cell.border = self.THIN_BORDER

                    col = self.pld_year_start
                    hours_cells = []
                    dollars_cells = []

                    for year in range(1, self.total_years + 1):
                        hours = labor_cat.get(f'year_{year}_hours', 0)
                        rate = labor_cat.get(f'year_{year}_rate', 0)

                        hours_col = col
                        cell = ws.cell(current_row, col)
                        cell.value = hours
                        cell.number_format = self.NUMBER_FORMAT
                        cell.border = self.THIN_BORDER
                        hours_cells.append(f"{get_column_letter(col)}{current_row}")
                        col += 1

                        rate_col = col
                        cell = ws.cell(current_row, col)
                        cell.value = rate
                        cell.number_format = self.CURRENCY_FORMAT
                        cell.border = self.THIN_BORDER
                        col += 1

                        cell = ws.cell(current_row, col)
                        cell.value = f"={get_column_letter(hours_col)}{current_row}*{get_column_letter(rate_col)}{current_row}"
                        cell.number_format = self.CURRENCY_FORMAT
                        cell.border = self.THIN_BORDER
                        dollars_cells.append(f"{get_column_letter(col)}{current_row}")
                        col += 1

                    cell = ws.cell(current_row, total_hours_col)
                    cell.value = f"={'+'.join(hours_cells)}" if hours_cells else 0
                    cell.number_format = self.NUMBER_FORMAT
                    cell.border = self.THIN_BORDER

                    cell = ws.cell(current_row, total_dollars_col)
                    cell.value = f"={'+'.join(dollars_cells)}" if dollars_cells else 0
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER

                    current_row += 1

        # Total row
        if positions or (self.has_gsa and any(sub.get('labor_categories') for sub in self.project_data.get('subcontractors', []))):
            ws.cell(current_row, 1, "Total Direct Labor")
            ws.cell(current_row, 1).font = self.BOLD_FONT
            ws.cell(current_row, 1).border = self.THIN_BORDER

            col = self.pld_year_start
            for year in range(1, self.total_years + 1):
                # Skip Hours column in sum (just dollars)
                col += 1  # Hours
                col += 1  # Rate

                # Dollars sum
                cell = ws.cell(current_row, col)
                cell.value = f"=SUM({get_column_letter(col)}{header_row + 1}:{get_column_letter(col)}{current_row - 1})"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                col += 1

            # Total Dollars sum
            cell = ws.cell(current_row, total_dollars_col)
            cell.value = f"=SUM({get_column_letter(total_dollars_col)}{header_row + 1}:{get_column_letter(total_dollars_col)}{current_row - 1})"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT

            total_direct_labor_row = current_row
            current_row += 1
            fringe_row = oh_row = ga_row = None  # only populated for BLS proposals

            if self.has_gsa:
                # GSA: fully-loaded rate stored directly — no Fringe/OH/G&A/Fee breakdown needed
                ws.cell(total_direct_labor_row, 1, "Total Prime Labor")
                total_prime_labor_row = total_direct_labor_row
            else:
                # BLS: Fringe / OH / G&A / Subtotal / Fee / Total Prime Labor rows
                fringe_row = current_row
                ws.cell(current_row, 1, "Fringe Benefits")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                for year in range(1, self.total_years + 1):
                    col += 1
                    rate_cell = ws.cell(current_row, col)
                    rate_cell.value = "='Indirect Rate'!B9"
                    rate_cell.number_format = self.PERCENT_FORMAT
                    rate_cell.border = self.THIN_BORDER
                    rate_col = col
                    col += 1
                    dl_cell = f"{get_column_letter(col)}{total_direct_labor_row}"
                    cell = ws.cell(current_row, col)
                    cell.value = f"={dl_cell}*{get_column_letter(rate_col)}{current_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    col += 1
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = f"={get_column_letter(total_dollars_col)}{total_direct_labor_row}*'Indirect Rate'!B9"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                current_row += 1

                oh_row = current_row
                ws.cell(current_row, 1, "Labor Overhead")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                for year in range(1, self.total_years + 1):
                    col += 1
                    rate_cell = ws.cell(current_row, col)
                    rate_cell.value = "='Indirect Rate'!B10"
                    rate_cell.number_format = self.PERCENT_FORMAT
                    rate_cell.border = self.THIN_BORDER
                    rate_col = col
                    col += 1
                    cl = get_column_letter(col)
                    cell = ws.cell(current_row, col)
                    cell.value = f"=({cl}{total_direct_labor_row}+{cl}{fringe_row})*{get_column_letter(rate_col)}{current_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    col += 1
                tcl = get_column_letter(total_dollars_col)
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = f"=({tcl}{total_direct_labor_row}+{tcl}{fringe_row})*'Indirect Rate'!B10"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                current_row += 1

                ga_row = current_row
                ws.cell(current_row, 1, "General & Administrative (Labor)")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                for year in range(1, self.total_years + 1):
                    col += 1
                    rate_cell = ws.cell(current_row, col)
                    rate_cell.value = "='Indirect Rate'!B12"
                    rate_cell.number_format = self.PERCENT_FORMAT
                    rate_cell.border = self.THIN_BORDER
                    rate_col = col
                    col += 1
                    cl = get_column_letter(col)
                    cell = ws.cell(current_row, col)
                    cell.value = f"=({cl}{total_direct_labor_row}+{cl}{fringe_row}+{cl}{oh_row})*{get_column_letter(rate_col)}{current_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    col += 1
                tcl = get_column_letter(total_dollars_col)
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = f"=({tcl}{total_direct_labor_row}+{tcl}{fringe_row}+{tcl}{oh_row})*'Indirect Rate'!B12"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                current_row += 1

                subtotal_row = current_row
                ws.cell(current_row, 1, "Subtotal")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                for year in range(1, self.total_years + 1):
                    col += 1
                    col += 1
                    cl = get_column_letter(col)
                    cell = ws.cell(current_row, col)
                    cell.value = f"={cl}{total_direct_labor_row}+{cl}{fringe_row}+{cl}{oh_row}+{cl}{ga_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    cell.font = self.BOLD_FONT
                    col += 1
                tcl = get_column_letter(total_dollars_col)
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = f"={tcl}{total_direct_labor_row}+{tcl}{fringe_row}+{tcl}{oh_row}+{tcl}{ga_row}"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                current_row += 1

                fee_row = current_row
                ws.cell(current_row, 1, "Fee")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                for year in range(1, self.total_years + 1):
                    col += 1
                    rate_cell = ws.cell(current_row, col)
                    rate_cell.value = "='Indirect Rate'!B14"
                    rate_cell.number_format = self.PERCENT_FORMAT
                    rate_cell.border = self.THIN_BORDER
                    rate_col = col
                    col += 1
                    cl = get_column_letter(col)
                    cell = ws.cell(current_row, col)
                    cell.value = f"={cl}{subtotal_row}*{get_column_letter(rate_col)}{current_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    col += 1
                tcl = get_column_letter(total_dollars_col)
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = f"={tcl}{subtotal_row}*'Indirect Rate'!B14"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                current_row += 1

                # ── Overtime row (OT hours × fee-inclusive FBLR × ot_multiplier) ──
                # Written as literal dollar amounts per year (not an Excel formula)
                # because PLD position rows don't have OT-hours columns.
                ot_by_year = self._compute_prime_ot_by_year()
                overtime_row = current_row
                ws.cell(current_row, 1, "Overtime (Prime)")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                ot_total = 0.0
                for year in range(1, self.total_years + 1):
                    col += 1  # hours column (unused for OT aggregate row)
                    col += 1  # rate column (unused)
                    amt = round(ot_by_year.get(year, 0.0), 2)
                    cell = ws.cell(current_row, col)
                    cell.value = amt
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    cell.font = self.BOLD_FONT
                    ot_total += amt
                    col += 1
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = round(ot_total, 2)
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                current_row += 1

                total_prime_labor_row = current_row
                ws.cell(current_row, 1, "Total Prime Labor")
                ws.cell(current_row, 1).font = self.BOLD_FONT
                ws.cell(current_row, 1).border = self.THIN_BORDER
                col = self.pld_year_start
                for year in range(1, self.total_years + 1):
                    col += 1
                    col += 1
                    cl = get_column_letter(col)
                    cell = ws.cell(current_row, col)
                    cell.value = f"={cl}{subtotal_row}+{cl}{fee_row}+{cl}{overtime_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    cell.font = self.BOLD_FONT
                    col += 1
                tcl = get_column_letter(total_dollars_col)
                cell = ws.cell(current_row, total_dollars_col)
                cell.value = f"={tcl}{subtotal_row}+{tcl}{fee_row}+{tcl}{overtime_row}"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT

        # Return row/col info so CE Summary and FLLR can reference this sheet
        return {
            'total_dl_row': total_direct_labor_row if positions else None,
            'fringe_row': fringe_row if (positions and not self.has_gsa) else None,
            'oh_row': oh_row if (positions and not self.has_gsa) else None,
            'ga_row': ga_row if (positions and not self.has_gsa) else None,
            'overtime_row': overtime_row if (positions and not self.has_gsa) else None,
            'total_prime_labor_row': total_prime_labor_row if positions else None,
            'total_dollars_col': total_dollars_col,
        }

    def _create_subcontractor_sheet(self, sub_data: Dict):
        """Create a sheet for a single subcontractor, named by company."""
        # Use company name for sheet title (sanitize for Excel sheet name limits)
        company_name = sub_data.get('name', 'Subcontractor')
        sheet_name = self._sanitize_sheet_name(company_name)
        ws = self.wb.create_sheet(sheet_name)

        # Column widths matching template
        # A: Labor Category (was B)
        ws.column_dimensions['A'].width = 50.5
        # B: Site (was C)
        ws.column_dimensions['B'].width = 19.66

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Column headers
        header_row = 10
        headers = ["Labor Category", "Site"]
        for idx, header in enumerate(headers):
            cell = ws.cell(header_row, 1 + idx)
            cell.value = header
            self._style_header_cell(cell)

        # Period columns
        col = 3
        for year in range(1, self.total_years + 1):
            period_label = self._get_period_label(year)

            # Period header above (row 7) - Light blue background, merged across 3 columns
            period_cell = ws.cell(7, col, period_label)
            self._style_period_header_cell(period_cell)
            # Merge across Hours, Rate, Dollars columns
            ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 2)

            # Sub-headers for Hours, Rate, Dollars
            sub_headers = ["Hours/Base", "Rate", "Dollars"]
            for sub_idx, sub_header in enumerate(sub_headers):
                cell = ws.cell(header_row, col + sub_idx)
                cell.value = sub_header
                self._style_header_cell(cell)

                # Set column widths matching template
                if sub_idx == 0:  # Hours
                    ws.column_dimensions[get_column_letter(col + sub_idx)].width = 14.33
                elif sub_idx == 1:  # Rate
                    ws.column_dimensions[get_column_letter(col + sub_idx)].width = 11.33
                else:  # Dollars
                    ws.column_dimensions[get_column_letter(col + sub_idx)].width = 14.5

            col += 3

        # Total column
        ws.cell(header_row, col, "Total")
        self._style_header_cell(ws.cell(header_row, col))
        total_col = col

        # Data rows
        current_row = header_row + 1
        for labor_cat in sub_data.get('labor_categories', []):
            ws.cell(current_row, 1, labor_cat['labor_category'])
            ws.cell(current_row, 1).border = self.THIN_BORDER

            ws.cell(current_row, 2, labor_cat.get('site', 'Government'))
            ws.cell(current_row, 2).border = self.THIN_BORDER

            col = 3
            dollars_cells = []

            for year in range(1, self.total_years + 1):
                hours = labor_cat.get(f'year_{year}_hours', 0)
                rate = labor_cat.get(f'year_{year}_rate', 0)

                # Hours
                cell = ws.cell(current_row, col)
                cell.value = hours
                cell.number_format = self.NUMBER_FORMAT
                cell.border = self.THIN_BORDER
                col += 1

                # Rate
                cell = ws.cell(current_row, col)
                cell.value = rate
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                hours_col = col - 1
                rate_col = col
                col += 1

                # Dollars (formula)
                cell = ws.cell(current_row, col)
                cell.value = f"={get_column_letter(hours_col)}{current_row}*{get_column_letter(rate_col)}{current_row}"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                dollars_cells.append(f"{get_column_letter(col)}{current_row}")
                col += 1

            # Total
            cell = ws.cell(current_row, total_col)
            cell.value = f"={'+'.join(dollars_cells)}"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT

            current_row += 1

        # Total row - sum hours and dollars by column (only if there are labor categories)
        if sub_data.get('labor_categories'):
            ws.cell(current_row, 1, "Total")
            ws.cell(current_row, 1).font = self.BOLD_FONT
            ws.cell(current_row, 1).border = self.THIN_BORDER

            col = 3
            for year in range(1, self.total_years + 1):
                # Hours column - sum
                hours_col_letter = get_column_letter(col)
                cell = ws.cell(current_row, col)
                cell.value = f"=SUM({hours_col_letter}{header_row + 1}:{hours_col_letter}{current_row - 1})"
                cell.number_format = self.NUMBER_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                col += 1

                # Rate column - skip (no total for rates)
                col += 1

                # Dollars column - sum
                dollars_col_letter = get_column_letter(col)
                cell = ws.cell(current_row, col)
                cell.value = f"=SUM({dollars_col_letter}{header_row + 1}:{dollars_col_letter}{current_row - 1})"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT
                col += 1

            # Total column - sum
            total_col_letter = get_column_letter(total_col)
            cell = ws.cell(current_row, total_col)
            cell.value = f"=SUM({total_col_letter}{header_row + 1}:{total_col_letter}{current_row - 1})"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT

        # Return info for CE Summary and FLLR cross-sheet references
        return {
            'sheet_name': sheet_name,
            'total_row': current_row,
            'total_col': total_col,
        }

    def _create_fully_loaded_labor_rates_sheet(self):
        """Create the Fully Loaded Labor Rates sheet — feeder sheet 2.

        All values are Excel formulas, not Python-computed numbers:
          BLS prime:  = 'Prime Labor Detail'!rate_col × IR burden cascade
                        (PLD year 1 rate itself feeds from BLS Analysis via INDEX/MATCH)
          GSA prime:  = 'Prime Labor Detail'!rate_col  (already fully loaded in PLD)
          Sub:        = '{sub_sheet}'!rate_col          (direct reference to sub sheet)

        PLD column layout: year N rate col = 3N+3  (data starts row 11, col D is GSA Discount)
        Sub column layout: year N rate col = 3N+1  (data starts row 11)
        """
        ws = self.wb.create_sheet("Fully Loaded Labor Rates")

        # Column widths
        ws.column_dimensions['A'].width = 50.5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Column headers
        header_row = 8
        ws.cell(header_row, 1, "Labor Category")
        self._style_header_cell(ws.cell(header_row, 1))
        ws.cell(header_row, 2, "Company")
        self._style_header_cell(ws.cell(header_row, 2))
        ws.cell(header_row, 3, "Location")
        self._style_header_cell(ws.cell(header_row, 3))

        col = 4
        for year in range(1, self.total_years + 1):
            cell = ws.cell(header_row, col)
            cell.value = self._get_period_label(year)
            self._style_period_header_cell(cell)
            ws.column_dimensions[get_column_letter(col)].width = 18
            col += 1

        # PLD and sub sheet data rows both start at row 11 (header_row = 10)
        PLD_DATA_START = 11
        SUB_DATA_START = 11

        current_row = header_row + 1

        # ── Prime positions ──────────────────────────────────────────────────
        for i, position in enumerate(self.project_data.get('prime_positions', [])):
            pld_row = PLD_DATA_START + i

            ws.cell(current_row, 1, position['labor_category'])
            ws.cell(current_row, 1).border = self.THIN_BORDER
            ws.cell(current_row, 2, self.project_data['prime_contractor_name'])
            ws.cell(current_row, 2).border = self.THIN_BORDER
            ws.cell(current_row, 3, position.get('location', ''))
            ws.cell(current_row, 3).border = self.THIN_BORDER

            col = 4
            for year in range(1, self.total_years + 1):
                # PLD rate column for year N = 3N+3 (col D is GSA Discount)
                pld_rate_col = get_column_letter(3 * year + self.pld_year_start - 2)
                pld_ref = f"'Prime Labor Detail'!{pld_rate_col}{pld_row}"

                cell = ws.cell(current_row, col)
                if (position.get('wage_source') or '').lower() == 'gsa':
                    # GSA rate in PLD is already fully loaded — reference directly
                    cell.value = f"={pld_ref}"
                else:
                    # BLS: FBLR = DL rate × (1+fringe) × (1+OH) × (1+G&A) × (1+fee)
                    location_type = position.get('location_type', 'On-Site')
                    oh_ir_row = self.IR_OH_ONSITE_ROW if location_type == 'On-Site' else self.IR_OH_OFFSITE_ROW
                    cell.value = (
                        f"={pld_ref}"
                        f"*(1+{self._ir_ref(self.IR_FRINGE_ROW)})"
                        f"*(1+{self._ir_ref(oh_ir_row)})"
                        f"*(1+{self._ir_ref(self.IR_GA_ROW)})"
                        f"*(1+{self._ir_ref(self.IR_FEE_LABOR_ROW)})"
                    )
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                col += 1

            current_row += 1

        # ── Subcontractor positions ───────────────────────────────────────────
        for sub in self.project_data.get('subcontractors', []):
            sub_name = sub.get('name', 'Subcontractor')
            sheet_name = self._sanitize_sheet_name(sub_name)

            for lc_idx, labor_cat in enumerate(sub.get('labor_categories', [])):
                sub_lc_row = SUB_DATA_START + lc_idx

                ws.cell(current_row, 1, labor_cat['labor_category'])
                ws.cell(current_row, 1).border = self.THIN_BORDER
                ws.cell(current_row, 2, sub_name)
                ws.cell(current_row, 2).border = self.THIN_BORDER
                ws.cell(current_row, 3, labor_cat.get('location', ''))
                ws.cell(current_row, 3).border = self.THIN_BORDER

                col = 4
                for year in range(1, self.total_years + 1):
                    # Sub rate column for year N = 3N+1 (sub starts at col 3)
                    sub_rate_col = get_column_letter(3 * year + 1)
                    cell = ws.cell(current_row, col)
                    cell.value = f"='{sheet_name}'!{sub_rate_col}{sub_lc_row}"
                    cell.number_format = self.CURRENCY_FORMAT
                    cell.border = self.THIN_BORDER
                    col += 1

                current_row += 1

    def _create_odcs_sheet(self):
        """Create the ODCs sheet (separate from Materials)."""
        ws = self.wb.create_sheet("ODCs")

        # A: ODC description
        ws.column_dimensions['A'].width = 40.66

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Period columns (single column per period - includes handling)
        header_row = 8
        col = 2

        for year in range(1, self.total_years + 1):
            period_label = self._get_period_label(year)
            period_cell = ws.cell(header_row, col, period_label)
            self._style_period_header_cell(period_cell)
            ws.column_dimensions[get_column_letter(col)].width = 18
            col += 1

        # Total column
        ws.cell(header_row, col, "Total")
        self._style_header_cell(ws.cell(header_row, col))
        total_col = col

        # ODC rows — base only (no passthrough baked in here; handling shown separately)
        current_row = header_row + 1
        odcs = self.project_data.get('odcs', [])
        odc_start_row = current_row

        for odc in odcs:
            ws.cell(current_row, 1, odc['category'])
            ws.cell(current_row, 1).border = self.THIN_BORDER
            escalate = odc.get('escalate', False)

            col = 2
            year1_base = odc.get('amount_year_1') or 0
            for year in range(1, self.total_years + 1):
                cell = ws.cell(current_row, col)
                if escalate and 'amount_per_year' not in odc:
                    # True feeder: year1 base × IR escalation formula (base only)
                    cell.value = f"={year1_base}{self._escalation_formula(1, year)}"
                else:
                    if 'amount_per_year' in odc:
                        base_amount = odc['amount_per_year'].get(str(year)) or 0
                    else:
                        base_amount = year1_base
                    cell.value = base_amount
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                col += 1

            total_cell = ws.cell(current_row, total_col)
            total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
            total_cell.number_format = self.CURRENCY_FORMAT
            total_cell.border = self.THIN_BORDER
            current_row += 1

        # Base ODCs subtotal row
        base_only_total_row = current_row
        ws.cell(current_row, 1, "Base ODCs")
        ws.cell(current_row, 1).font = self.BOLD_FONT
        ws.cell(current_row, 1).border = self.THIN_BORDER
        col = 2
        for year in range(1, self.total_years + 1):
            cell = ws.cell(current_row, col)
            cell.value = f"=SUM({get_column_letter(col)}{odc_start_row}:{get_column_letter(col)}{current_row - 1})"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 1
        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        total_cell.font = self.BOLD_FONT
        current_row += 1

        # Materials Handling (S&MH) row = base_only_total × S&MH-only rate
        # (IR_SMH_MATERIALS_ROW, not IR_PASSTHROUGH_ROW — materials don't get G&A passthrough;
        # matches the Nexagen sample template's Material Handling formula.)
        handling_row = current_row
        ws.cell(current_row, 1, "Materials Handling (S&MH)")
        ws.cell(current_row, 1).border = self.THIN_BORDER
        col = 2
        for year in range(1, self.total_years + 1):
            cell = ws.cell(current_row, col)
            cell.value = f"={get_column_letter(col)}{base_only_total_row}*{self._ir_ref(self.IR_SMH_MATERIALS_ROW)}"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            col += 1
        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        current_row += 1

        # Total ODCs row = Base + Handling
        ws.cell(current_row, 1, "Total ODCs")
        ws.cell(current_row, 1).font = self.BOLD_FONT
        ws.cell(current_row, 1).border = self.THIN_BORDER
        col = 2
        for year in range(1, self.total_years + 1):
            cl = get_column_letter(col)
            cell = ws.cell(current_row, col)
            cell.value = f"={cl}{base_only_total_row}+{cl}{handling_row}"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 1
        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        total_cell.font = self.BOLD_FONT

        return {
            'base_only_total_row': base_only_total_row,   # Base materials, no handling
            'handling_row': handling_row,                  # S&MH row
            'total_col': total_col,
        }

    def _create_materials_sheet(self):
        """Create the Materials sheet (separate from ODCs)."""
        ws = self.wb.create_sheet("Materials")

        # B: Material description (was B, now A)
        ws.column_dimensions['A'].width = 40.66

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Period columns
        header_row = 8
        col = 2

        for year in range(1, self.total_years + 1):
            period_label = self._get_period_label(year)
            # Merge period label across the 2 columns (base + handling)
            period_cell = ws.cell(header_row, col, period_label)
            self._style_period_header_cell(period_cell)
            ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 1)
            ws.column_dimensions[get_column_letter(col)].width = 18
            ws.column_dimensions[get_column_letter(col + 1)].width = 15
            col += 2

        # Total column
        ws.cell(header_row, col, "Total")
        self._style_header_cell(ws.cell(header_row, col))
        total_col = col

        # Material rows
        current_row = header_row + 2
        materials = self.project_data.get('materials', [])
        material_start_row = current_row

        for material in materials:
            # Material base row
            ws.cell(current_row, 1, material['category'])
            ws.cell(current_row, 1).border = self.THIN_BORDER
            escalate = material.get('escalate', False)

            col = 2
            year1_base = material.get('amount_year_1') or 0
            for year in range(1, self.total_years + 1):
                cell = ws.cell(current_row, col)
                if escalate and 'amount_per_year' not in material:
                    # True feeder: year1 base × IR escalation formula
                    cell.value = f"={year1_base}{self._escalation_formula(1, year)}"
                else:
                    # Pre-set amounts per year or no escalation: Python value
                    if 'amount_per_year' in material:
                        base_amount = material['amount_per_year'].get(str(year)) or 0
                    else:
                        base_amount = year1_base
                    cell.value = base_amount
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                col += 2

            # Add Total column formula for material row
            total_cell = ws.cell(current_row, total_col)
            total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 2)}{current_row})"
            total_cell.number_format = self.CURRENCY_FORMAT
            total_cell.border = self.THIN_BORDER

            current_row += 1

            # Material Handling row — uses S&MH-only rate (not combined passthrough)
            ws.cell(current_row, 1, f"{material['category']} Handling")
            ws.cell(current_row, 1).border = self.THIN_BORDER

            col = 2
            for year in range(1, self.total_years + 1):
                material_cell = f"{get_column_letter(col)}{current_row - 1}"
                cell = ws.cell(current_row, col)
                cell.value = f"={material_cell}*{self._ir_ref(self.IR_SMH_MATERIALS_ROW)}"
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                col += 2

            # Add Total column formula for handling row
            total_cell = ws.cell(current_row, total_col)
            total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 2)}{current_row})"
            total_cell.number_format = self.CURRENCY_FORMAT
            total_cell.border = self.THIN_BORDER

            current_row += 1

        # Total row
        ws.cell(current_row, 1, "Total Materials")
        ws.cell(current_row, 1).font = self.BOLD_FONT
        ws.cell(current_row, 1).border = self.THIN_BORDER

        col = 2
        for year in range(1, self.total_years + 1):
            cell = ws.cell(current_row, col)
            cell.value = f"=SUM({get_column_letter(col)}{material_start_row}:{get_column_letter(col)}{current_row - 1})"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 2

        # Add Total column formula for Total Materials row
        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 2)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        total_cell.font = self.BOLD_FONT

    def _create_travel_sheet(self):
        """Create the Travel sheet."""
        ws = self.wb.create_sheet("Travel")

        # A: Travel description
        ws.column_dimensions['A'].width = 40.66

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Period columns (single column per period - includes G&A)
        header_row = 8
        col = 2

        for year in range(1, self.total_years + 1):
            period_label = self._get_period_label(year)
            period_cell = ws.cell(header_row, col, period_label)
            self._style_period_header_cell(period_cell)  # Light blue background for periods
            ws.column_dimensions[get_column_letter(col)].width = 18
            col += 1

        # Total column
        ws.cell(header_row, col, "Total")
        self._style_header_cell(ws.cell(header_row, col))
        total_col = col

        # Travel rows
        current_row = header_row + 1
        travel_items = self.project_data.get('travel', [])
        travel_start_row = current_row

        for travel in travel_items:
            description = travel.get('description', 'Travel')
            escalate = travel.get('escalate', False)

            # Travel row — raw base amount (no G&A); G&A shown separately below
            ws.cell(current_row, 1, description)
            ws.cell(current_row, 1).border = self.THIN_BORDER

            col = 2
            year1_base = travel.get('amount_year_1') or 0
            for year in range(1, self.total_years + 1):
                cell = ws.cell(current_row, col)
                if escalate and 'amount_per_year' not in travel:
                    # True feeder: year1 base × IR escalation formula (raw, no G&A)
                    cell.value = f"={year1_base}{self._escalation_formula(1, year)}"
                else:
                    # Pre-set amounts per year or no escalation: raw amount
                    if 'amount_per_year' in travel:
                        base_amount = travel['amount_per_year'].get(str(year)) or 0
                    else:
                        base_amount = year1_base
                    cell.value = base_amount
                cell.number_format = self.CURRENCY_FORMAT
                cell.border = self.THIN_BORDER
                col += 1

            # Total column for this item row
            total_cell = ws.cell(current_row, total_col)
            total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
            total_cell.number_format = self.CURRENCY_FORMAT
            total_cell.border = self.THIN_BORDER

            current_row += 1

        # Base Travel subtotal row (raw, no G&A) — referenced by CE Summary row 18
        base_total_row = current_row
        ws.cell(current_row, 1, "Base Travel")
        ws.cell(current_row, 1).font = self.BOLD_FONT
        ws.cell(current_row, 1).border = self.THIN_BORDER

        col = 2
        for year in range(1, self.total_years + 1):
            cell = ws.cell(current_row, col)
            cell.value = f"=SUM({get_column_letter(col)}{travel_start_row}:{get_column_letter(col)}{current_row - 1})"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 1

        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        total_cell.font = self.BOLD_FONT
        current_row += 1

        # G&A on Travel row (display only — CE Summary computes its own G&A from Base Travel)
        ga_row = current_row
        ws.cell(current_row, 1, "G&A on Travel")
        ws.cell(current_row, 1).border = self.THIN_BORDER

        col = 2
        for year in range(1, self.total_years + 1):
            cell = ws.cell(current_row, col)
            cell.value = f"={get_column_letter(col)}{base_total_row}*{self._ir_ref(self.IR_GA_ROW)}"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            col += 1

        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        current_row += 1

        # Total Travel (Base + G&A) — for display in the Travel sheet
        ws.cell(current_row, 1, "Total Travel")
        ws.cell(current_row, 1).font = self.BOLD_FONT
        ws.cell(current_row, 1).border = self.THIN_BORDER

        col = 2
        for year in range(1, self.total_years + 1):
            cell = ws.cell(current_row, col)
            cell.value = f"={get_column_letter(col)}{base_total_row}+{get_column_letter(col)}{ga_row}"
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 1

        total_cell = ws.cell(current_row, total_col)
        total_cell.value = f"=SUM({get_column_letter(2)}{current_row}:{get_column_letter(total_col - 1)}{current_row})"
        total_cell.number_format = self.CURRENCY_FORMAT
        total_cell.border = self.THIN_BORDER
        total_cell.font = self.BOLD_FONT

        return {
            'base_total_row': base_total_row,
            'ga_on_travel_row': ga_row,        # G&A on Travel row — used by CE Summary separately
            'total_travel_row': current_row,   # Total Travel (Base + G&A) — combined
            'total_col': total_col,
        }

    def _create_loe_sheet(self):
        """Create the Level of Effort (LOE) sheet showing hours per category."""
        ws = self.wb.create_sheet("LOE")

        # Column widths matching template
        # A: Labor Category (was B)
        ws.column_dimensions['A'].width = 50.5
        # B: Company (was C)
        ws.column_dimensions['B'].width = 20
        # C: Site (was D)
        ws.column_dimensions['C'].width = 19.66
        # D: Location (was E)
        ws.column_dimensions['D'].width = 14.33

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Column headers
        header_row = 10
        headers = ["Labor Category", "Company", "Site", "Location"]
        for idx, header in enumerate(headers):
            cell = ws.cell(header_row, 1 + idx)
            cell.value = header
            self._style_header_cell(cell)

        # Period columns (Hours only)
        col = 5
        for year in range(1, self.total_years + 1):
            period_label = self._get_period_label(year)

            # Period header above (row 7) - Light blue background
            period_cell = ws.cell(7, col, period_label)
            self._style_period_header_cell(period_cell)

            # Column header for hours
            cell = ws.cell(header_row, col)
            cell.value = "Hours/Base"
            self._style_header_cell(cell)
            ws.column_dimensions[get_column_letter(col)].width = 14.33
            col += 1
            
        # Total Hours column
        ws.cell(header_row, col, "Total Hours")
        self._style_header_cell(ws.cell(header_row, col))
        total_col = col
        
        # Data rows
        current_row = header_row + 1
        positions = self.project_data.get('prime_positions', [])
        
        for position in positions:
            # Labor Category
            ws.cell(current_row, 1, position['labor_category'])
            ws.cell(current_row, 1).border = self.THIN_BORDER

            # Company (Prime)
            ws.cell(current_row, 2, self.project_data['prime_contractor_name'])
            ws.cell(current_row, 2).border = self.THIN_BORDER

            # Site
            ws.cell(current_row, 3, position.get('site', 'Government'))
            ws.cell(current_row, 3).border = self.THIN_BORDER

            # Location
            ws.cell(current_row, 4, position.get('location', ''))
            ws.cell(current_row, 4).border = self.THIN_BORDER

            # Hours per year
            hour_cells = []
            col = 5
            hours_data = position.get('hours_per_year', {})
            
            for year in range(1, self.total_years + 1):
                # Handle dictionary format {"1": 1880} or parsed format if different
                if isinstance(hours_data, dict):
                    hours = hours_data.get(str(year), 0)
                else:
                    hours = 0
                    
                cell = ws.cell(current_row, col)
                cell.value = hours
                cell.number_format = self.NUMBER_FORMAT
                cell.border = self.THIN_BORDER
                hour_cells.append(f"{get_column_letter(col)}{current_row}")
                col += 1
                
            # Total Hours
            cell = ws.cell(current_row, total_col)
            cell.value = f"={'+'.join(hour_cells)}"
            cell.number_format = self.NUMBER_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            
            current_row += 1

        # Add subcontractor positions
        for sub in self.project_data.get('subcontractors', []):
            sub_name = sub.get('name', 'Subcontractor')

            for labor_cat in sub.get('labor_categories', []):
                # Labor Category
                ws.cell(current_row, 1, labor_cat['labor_category'])
                ws.cell(current_row, 1).border = self.THIN_BORDER

                # Company (Subcontractor)
                ws.cell(current_row, 2, sub_name)
                ws.cell(current_row, 2).border = self.THIN_BORDER

                # Site
                ws.cell(current_row, 3, labor_cat.get('site', 'Government'))
                ws.cell(current_row, 3).border = self.THIN_BORDER

                # Location
                ws.cell(current_row, 4, labor_cat.get('location', ''))
                ws.cell(current_row, 4).border = self.THIN_BORDER

                # Hours per year
                hour_cells = []
                col = 5
                for year in range(1, self.total_years + 1):
                    hours = labor_cat.get(f'year_{year}_hours', 0)

                    cell = ws.cell(current_row, col)
                    cell.value = hours
                    cell.number_format = self.NUMBER_FORMAT
                    cell.border = self.THIN_BORDER
                    hour_cells.append(f"{get_column_letter(col)}{current_row}")
                    col += 1

                # Total Hours
                cell = ws.cell(current_row, total_col)
                cell.value = f"={'+'.join(hour_cells)}"
                cell.number_format = self.NUMBER_FORMAT
                cell.border = self.THIN_BORDER
                cell.font = self.BOLD_FONT

                current_row += 1

        # Total Row (for all positions - prime + subcontractors)
        ws.cell(current_row, 1, "Total")
        ws.cell(current_row, 1).font = self.BOLD_FONT
        ws.cell(current_row, 1).border = self.THIN_BORDER

        col = 5
        for year in range(1, self.total_years + 2): # Periods + Total column
            cell = ws.cell(current_row, col)
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{header_row+1}:{col_letter}{current_row-1})"
            cell.number_format = self.NUMBER_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 1

    def _create_indirect_rates_sheet(self):
        """Create the Indirect Rates reference sheet."""
        ws = self.wb.create_sheet("Indirect Rate")

        # Column widths matching template
        # A: Rate description (was B)
        ws.column_dimensions['A'].width = 40.66
        # B: Rate value (was C)
        ws.column_dimensions['B'].width = 15

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Indirect Rates header (with blue background like template)
        header_cell = ws.cell(8, 1, "Indirect Rates")
        header_cell.font = self.HEADER_FONT
        header_cell.fill = self.HEADER_FILL
        header_cell.border = self.THIN_BORDER
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A8:B8')

        # Rate rows
        indirect_rates = self.project_data.get('indirect_rates', {})
        passthrough_rates = self.project_data.get('passthrough_rates', {})
        fee_rates = self.project_data.get('fee_rates', {})
        escalation_rates = self.project_data.get('escalation_rates', {})

        # Passthrough (subcontractor) = S&MH + G&A Passthrough (combined).
        # Material Handling (ODC) = S&MH only (matches Nexagen sample template).
        smh_only = passthrough_rates.get('smh', 0)
        combined_passthrough = smh_only + passthrough_rates.get('ga', 0)

        # For GSA only G&A is used in formulas; Fringe/OH/Passthrough/Fee are not applied.
        # Row positions must stay fixed (IR_*_ROW constants are hardcoded across all sheets).
        # For GSA: hide unused rows so they don't clutter the sheet.
        all_rates = [
            ("Fringe", indirect_rates.get('fringe', 0)),
            ("Onsite Overhead (OH)", indirect_rates.get('oh_onsite', indirect_rates.get('oh', 0))),
            ("Offsite Overhead (OH)", indirect_rates.get('oh_offsite', indirect_rates.get('oh', 0))),
            ("General & Administrative (G&A)", indirect_rates.get('ga', 0)),
            ("Passthrough (S&MH + G&A)", combined_passthrough),
            ("Fee on Labor", fee_rates.get('prime_labor', 0)),
            ("Fee on Subcontractor", fee_rates.get('sub_labor', 0)),
            ("Material Handling (S&MH)", smh_only),
        ]
        # Index 3 = G&A = IR_GA_ROW (row 12) — the only rate used in GSA formulas
        gsa_visible_indices = {3}

        current_row = 9
        for idx, (label, rate) in enumerate(all_rates):
            if self.has_gsa and idx not in gsa_visible_indices:
                ws.row_dimensions[current_row].hidden = True
                current_row += 1
                continue

            label_cell = ws.cell(current_row, 1, label)
            label_cell.font = self.BOLD_FONT
            label_cell.border = self.THIN_BORDER

            cell = ws.cell(current_row, 2)
            cell.value = rate
            cell.number_format = self.PERCENT_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.NORMAL_FONT

            current_row += 1

        # Escalation factors (start row 17 by constant; the rate loop above
        # ended at row 17 since we have 8 rate rows starting at row 9).
        for year in range(1, self.total_years):
            key = f"{year}_to_{year + 1}"
            rate = escalation_rates.get(key, 0.0)

            label = f"Escalation Factor (Year {year} to {year + 1})"
            label_cell = ws.cell(current_row, 1, label)
            label_cell.font = self.BOLD_FONT
            label_cell.border = self.THIN_BORDER

            cell = ws.cell(current_row, 2)
            cell.value = rate
            cell.number_format = self.PERCENT_FORMAT
            cell.border = self.THIN_BORDER
            cell.font = self.NORMAL_FONT

            current_row += 1

        # Surge Multiplier — only add if proposal has a surge option
        if self.surge_percentage:
            current_row += 1  # blank separator
            label_cell = ws.cell(current_row, 1, "Surge Multiplier")
            label_cell.font = self.BOLD_FONT
            label_cell.border = self.THIN_BORDER
            cell = ws.cell(current_row, 2)
            cell.value = self.surge_multiplier
            cell.number_format = '0.00'
            cell.border = self.THIN_BORDER
            cell.font = self.NORMAL_FONT
            self.ir_surge_multiplier_row = current_row

    def _calculate_cost_elements(self) -> Dict[str, Dict[str, float]]:
        """
        Calculate all cost elements for the CE Summary sheet.

        Returns a dict with keys for each cost element, each containing
        year values and total.
        """
        result = {
            'direct_labor': {},
            'fringe': {},
            'overhead': {},
            'ga_labor': {},
            'subcontractors': {},
            'sub_handling': {},
            'materials': {},
            'material_handling': {},
            'travel': {},
            'ga_travel': {},
            'fee': {},
        }

        indirect_rates = self.project_data.get('indirect_rates', {})
        fringe_rate = indirect_rates.get('fringe', 0.25)
        oh_onsite_rate = indirect_rates.get('oh_onsite', indirect_rates.get('oh', 0.07))
        oh_offsite_rate = indirect_rates.get('oh_offsite', indirect_rates.get('oh', 0.07))
        ga_rate = indirect_rates.get('ga', 0.22)
        smh_rate = self.project_data.get('passthrough_rates', {}).get('smh', 0.0671)
        fee_rates = self.project_data.get('fee_rates', {})
        prime_fee_rate = fee_rates.get('prime_labor', 0.08)
        sub_fee_rate = fee_rates.get('sub_labor', 0.0)

        # Calculate prime labor costs
        for year in range(1, self.total_years + 1):
            year_key = f"year_{year}"

            dl_total = 0
            fringe_total = 0
            oh_total = 0

            for position in self.project_data.get('prime_positions', []):
                # Use appropriate calculator based on wage source
                if (position.get('wage_source') or '').lower() == 'gsa':
                    # GSA positions: Calculate with GSA-specific function
                    discount_rate = position.get('gsa_discount_rate', 0.0)
                    results = Calculator.calculate_gsa_position_years(
                        position_data=position,
                        total_years=self.total_years,
                        discount_rate=discount_rate,
                        escalation_rates=self.project_data.get('escalation_rates', {})
                    )
                    year_data = results.get(year_key, {})

                    # For GSA: Reverse engineer DL from the fully loaded rate for CE Summary display.
                    # GSA rate is FULLY LOADED (includes fee), so we reverse-engineer with fee in multiplier.
                    # Honor location_type so the displayed OH slice matches the frontend's
                    # reverseEngineerGSARate (see PRICING_FORMULAS.md § 6.4).
                    gsa_rate = year_data.get('rate', 0)
                    hours = year_data.get('hours', 0)
                    position_location_type = position.get('location_type', 'On-Site')
                    position_oh_rate = oh_onsite_rate if position_location_type == 'On-Site' else oh_offsite_rate

                    multiplier = (1 + fringe_rate) * (1 + position_oh_rate) * (1 + ga_rate) * (1 + prime_fee_rate)
                    dl_rate = gsa_rate / multiplier if multiplier > 0 else 0
                    position_dl = dl_rate * hours
                    position_fringe = position_dl * fringe_rate
                    position_oh = (position_dl + position_fringe) * position_oh_rate
                else:
                    # BLS positions: Use standard calculator (respects location_type via indirect_rates)
                    results = Calculator.calculate_position_years(
                        position_data=position,
                        escalation_rates=self.project_data['escalation_rates'],
                        indirect_rates=indirect_rates,
                        total_years=self.total_years
                    )
                    year_data = results.get(year_key, {})

                    # Direct Labor = DL rate × hours
                    position_dl = year_data.get('dl_rate', 0) * year_data.get('hours', 0)
                    position_fringe = position_dl * fringe_rate

                    # For BLS: Use correct OH rate based on location_type
                    location_type = position.get('location_type', 'On-Site')
                    position_oh_rate = oh_onsite_rate if location_type == 'On-Site' else oh_offsite_rate
                    position_oh = (position_dl + position_fringe) * position_oh_rate

                # Accumulate totals
                dl_total += position_dl
                fringe_total += position_fringe
                oh_total += position_oh

            # Use full precision - no rounding to match UI exactly
            result['direct_labor'][year_key] = dl_total
            result['fringe'][year_key] = fringe_total
            result['overhead'][year_key] = oh_total
            subtotal2 = dl_total + fringe_total + oh_total
            result['ga_labor'][year_key] = subtotal2 * ga_rate

            # Prime labor total for fee calculation
            prime_labor_total = subtotal2 + result['ga_labor'][year_key]

            # Subcontractor costs
            sub_total = 0
            for sub in self.project_data.get('subcontractors', []):
                for labor_cat in sub.get('labor_categories', []):
                    hours = labor_cat.get(f'year_{year}_hours', 0)
                    rate = labor_cat.get(f'year_{year}_rate', 0)
                    sub_total += hours * rate

            result['subcontractors'][year_key] = sub_total
            result['sub_handling'][year_key] = sub_total * smh_rate

            # Materials with escalation
            material_total = 0
            escalation_rates = self.project_data.get('escalation_rates', {})
            for odc in self.project_data.get('odcs', []):
                if 'amount_per_year' in odc:
                    base_amount = odc['amount_per_year'].get(str(year)) or 0
                else:
                    base_amount = odc.get('amount_year_1') or 0

                # Apply compound escalation if flag is set
                escalated_amount = base_amount
                escalate = odc.get('escalate', False)
                if escalate and year > 1:
                    for y in range(1, year):
                        esc_key = f"{y}_to_{y + 1}"
                        esc_rate = escalation_rates.get(esc_key) or 0
                        escalated_amount *= (1 + esc_rate)

                material_total += escalated_amount

            result['materials'][year_key] = material_total
            result['material_handling'][year_key] = material_total * smh_rate

            # Travel with escalation
            travel_total = 0
            for travel in self.project_data.get('travel', []):
                if 'amount_per_year' in travel:
                    base_amount = travel['amount_per_year'].get(str(year)) or 0
                else:
                    base_amount = travel.get('amount_year_1') or 0

                # Apply compound escalation if flag is set
                escalated_amount = base_amount
                escalate = travel.get('escalate', False)
                if escalate and year > 1:
                    for y in range(1, year):
                        esc_key = f"{y}_to_{y + 1}"
                        esc_rate = escalation_rates.get(esc_key) or 0
                        escalated_amount *= (1 + esc_rate)

                travel_total += escalated_amount

            result['travel'][year_key] = travel_total
            result['ga_travel'][year_key] = travel_total * ga_rate

            # Fee (on labor only) - use full precision
            prime_fee = prime_labor_total * prime_fee_rate
            sub_labor_with_handling = sub_total + result['sub_handling'][year_key]
            sub_fee = sub_labor_with_handling * sub_fee_rate
            result['fee'][year_key] = prime_fee + sub_fee

        # Calculate totals
        for key in result:
            result[key]['total'] = sum(
                v for k, v in result[key].items() if k.startswith('year_')
            )

        return result

    def _style_header_cell(self, cell):
        """Apply header styling to a cell."""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.border = self.THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _style_period_header_cell(self, cell):
        """Apply period header styling to a cell (light blue background)."""
        cell.fill = self.PERIOD_HEADER_FILL
        cell.font = self.BOLD_FONT
        cell.border = self.THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _style_subheader_cell(self, cell):
        """Apply sub-header styling to a cell."""
        cell.font = self.BOLD_FONT
        cell.border = self.THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _create_bls_analysis_sheet(self):
        """Create the BLS Analysis sheet showing all positions with wage percentiles."""
        ws = self.wb.create_sheet(self.analysis_sheet_name)

        # Column widths
        ws.column_dimensions['A'].width = 30  # Labor Category
        ws.column_dimensions['B'].width = 20  # Location
        ws.column_dimensions['C'].width = 50  # Description
        ws.column_dimensions['D'].width = 10  # Source
        if self.has_gsa:
            # No SOC Code col — SOC Title shifts to E, Selected Wage at K
            ws.column_dimensions['E'].width = 35  # SOC Title / GSA Labor Category
            # Hide empty gap columns F–J (old percentile slots)
            for col_letter in ['F', 'G', 'H', 'I', 'J']:
                ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions['K'].width = 18  # Selected Wage/Rate
        else:
            ws.column_dimensions['E'].width = 15  # SOC Code
            ws.column_dimensions['F'].width = 35  # SOC Title
            ws.column_dimensions['G'].width = 15  # 10th Percentile
            ws.column_dimensions['H'].width = 15  # 25th Percentile
            ws.column_dimensions['I'].width = 15  # 50th Percentile
            ws.column_dimensions['J'].width = 15  # 75th Percentile
            ws.column_dimensions['K'].width = 15  # 90th Percentile
            ws.column_dimensions['L'].width = 18  # Selected Wage/Rate

        # Apply standard header format (Rows 1-5)
        self._apply_standard_header(ws, start_col=1)

        # Additional title row
        title_text = "Wage Data - All Positions" if self.has_gsa else "Wage Data - All Positions with Percentiles"
        ws.cell(7, 1, title_text)
        ws.cell(7, 1).font = self.BOLD_FONT

        # Column headers
        header_row = 8

        # Fixed headers A–D (always present)
        for idx, header in enumerate(["Labor Category", "Location", "Description", "Source"]):
            cell = ws.cell(header_row, 1 + idx)
            cell.value = header
            self._style_header_cell(cell)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if self.has_gsa:
            # GSA: no SOC Code col — SOC Title at col 5 (E), Selected Wage at col 11 (K)
            cell = ws.cell(header_row, 5, "SOC Title / GSA Labor Category")
            self._style_header_cell(cell)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            cell = ws.cell(header_row, 11, "Selected Wage/Rate")
            self._style_header_cell(cell)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # GSA Rate Year 2..N at cols 12, 13, ... (L, M, ...)
            for yr in range(2, self.total_years + 1):
                yr_col_idx = 11 + yr - 1
                cell = ws.cell(header_row, yr_col_idx, f"GSA Rate\nYear {yr}")
                self._style_header_cell(cell)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.column_dimensions[get_column_letter(yr_col_idx)].width = 14
        else:
            # BLS: SOC Code at col 5 (E), SOC Title at col 6 (F), percentiles G–K, Selected Wage at L
            for idx, header in enumerate(["SOC Code", "SOC Title / GSA Labor Category"]):
                cell = ws.cell(header_row, 5 + idx)
                cell.value = header
                self._style_header_cell(cell)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for idx, header in enumerate([
                "10th Percentile", "25th Percentile", "50th Percentile\n(Median)",
                "75th Percentile", "90th Percentile"
            ]):
                cell = ws.cell(header_row, 7 + idx)
                cell.value = header
                self._style_header_cell(cell)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            cell = ws.cell(header_row, 12, "Selected Wage/Rate")
            self._style_header_cell(cell)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows
        current_row = header_row + 1
        wage_data = self.project_data.get('wage_data', {})
        positions = wage_data.get('positions', [])

        for pos in positions:
            col = 1

            # Labor Category
            cell = ws.cell(current_row, col, pos.get('labor_category', ''))
            cell.border = self.THIN_BORDER
            cell.font = self.BOLD_FONT
            col += 1

            # Location
            cell = ws.cell(current_row, col, pos.get('location', ''))
            cell.border = self.THIN_BORDER
            col += 1

            # Description
            cell = ws.cell(current_row, col, pos.get('description', ''))
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1

            # Source (BLS or GSA)
            wage_source = pos.get('wage_source', 'bls').upper()
            cell = ws.cell(current_row, col, wage_source)
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if wage_source == 'GSA':
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                cell.font = Font(bold=True, color="1E40AF")
            else:
                cell.fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
                cell.font = Font(bold=True, color="7C3AED")
            col += 1

            # SOC Code (BLS proposals only — GSA proposals skip this column entirely)
            if not self.has_gsa:
                cell = ws.cell(current_row, col, format_soc_code(pos.get('soc_code', '')))
                cell.border = self.THIN_BORDER
                col += 1

            # SOC Title / GSA Labor Category
            if wage_source == 'GSA':
                cell = ws.cell(current_row, col, pos.get('gsa_title', ''))
            else:
                cell = ws.cell(current_row, col, pos.get('soc_title', ''))
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1

            # Percentiles (BLS only, GSA shows '-')
            # Strip " (default)" suffix from percentile field
            raw_percentile = pos.get('percentile', '50th')
            selected_percentile = raw_percentile.replace(' (default)', '') if raw_percentile else '50th'

            # For highlighting, check if user manually edited (selected_salaries exists)
            selected_salaries = pos.get('selected_salaries', [])
            highlighted_percentile = None
            if selected_salaries and len(selected_salaries) > 0:
                # User edited - determine which percentile matches the selected wage
                avg_wage = sum(selected_salaries) / len(selected_salaries)
                rounded_avg = round(avg_wage)
                # Check which percentile wage matches
                for pct in ['10th', '25th', '50th', '75th', '90th']:
                    pct_wage = pos.get(f'wage_{pct}')
                    if pct_wage and round(pct_wage) == rounded_avg:
                        highlighted_percentile = pct
                        break
            else:
                # No user edit - use system's selected percentile
                highlighted_percentile = selected_percentile

            # Only show percentile columns for BLS positions (skip for GSA)
            if wage_source == 'BLS':
                for percentile in ['10th', '25th', '50th', '75th', '90th']:
                    wage_key = f'wage_{percentile}'
                    wage_value = pos.get(wage_key)
                    cell = ws.cell(current_row, col)

                    if wage_value is not None:
                        cell.value = wage_value
                        cell.number_format = self.CURRENCY_FORMAT

                        # Highlight the percentile that's actually selected
                        if percentile == highlighted_percentile:
                            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                            cell.font = Font(bold=True, color="059669")
                        else:
                            cell.font = Font(color="7C3AED")
                    else:
                        cell.value = '-'
                        cell.alignment = Alignment(horizontal='center')

                    cell.border = self.THIN_BORDER
                    col += 1
            else:
                # GSA: Skip percentile columns entirely (leave blank)
                col += 5  # Skip 5 percentile columns

            # Selected Wage/Rate
            cell = ws.cell(current_row, col)

            if wage_source == 'GSA':
                # For GSA, get the current year's rate
                gsa_rates = pos.get('gsa_rates_by_year', {})
                current_year = pos.get('gsa_current_year', 1)
                custom_rate = pos.get('gsa_custom_rate')

                if custom_rate is not None:
                    selected_wage = custom_rate
                else:
                    selected_wage = gsa_rates.get(str(current_year), 0)
            else:
                # For BLS: prioritize user edits, then system selection
                if selected_salaries and len(selected_salaries) > 0:
                    # User edited - show average
                    selected_wage = sum(selected_salaries) / len(selected_salaries)
                elif pos.get('selected_wage'):
                    # System selected wage
                    selected_wage = pos.get('selected_wage')
                else:
                    # Fallback to percentile lookup (with cleaned percentile)
                    selected_wage = pos.get(f'wage_{selected_percentile}', 0)

            cell.value = selected_wage if selected_wage else 0
            cell.number_format = self.CURRENCY_FORMAT
            cell.border = self.THIN_BORDER
            cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            cell.font = Font(bold=True, color="1E40AF")

            # GSA year 2..N catalog rates — only for GSA proposals
            if self.has_gsa and wage_source == 'GSA':
                gsa_rates = pos.get('gsa_rates_by_year', {})
                gsa_current_year = pos.get('gsa_current_year', 1)
                for yr in range(2, self.total_years + 1):
                    yr_col_idx   = 11 + yr - 1          # K=11(yr1), L=12(yr2), M=13(yr3)...
                    prev_col     = get_column_letter(yr_col_idx - 1)
                    yr_cell      = ws.cell(current_row, yr_col_idx)
                    schedule_yr  = gsa_current_year + yr - 1
                    actual_rate  = gsa_rates.get(str(schedule_yr))
                    if actual_rate is not None:
                        yr_cell.value = actual_rate
                    else:
                        esc_row = self.IR_ESCALATION_START_ROW + yr - 2
                        yr_cell.value = f"={prev_col}{current_row}*(1+'Indirect Rate'!$B${esc_row})"
                    yr_cell.number_format = self.CURRENCY_FORMAT
                    yr_cell.border = self.THIN_BORDER

            current_row += 1

        # Set row height for better readability
        for row_idx in range(header_row + 1, current_row):
            ws.row_dimensions[row_idx].height = 40
