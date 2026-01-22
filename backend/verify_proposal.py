"""
Verify proposal data and compare with exported Excel.
"""
import openpyxl
from pymongo import MongoClient
from bson.objectid import ObjectId
from app.settings import settings

# Proposal to verify
PROPOSAL_ID = "6971506b72ec4b0d0502c3b2"
EXCEL_PATH = "/Users/keshav/Downloads/PriceIQ Pricing.xlsx"


def fetch_proposal_from_mongo():
    """Fetch proposal data from MongoDB."""
    client = MongoClient(settings.MONGODB_URL)
    db = client[settings.MONGODB_DATABASE]

    proposal = db.proposals.find_one({"_id": ObjectId(PROPOSAL_ID)})

    if not proposal:
        print(f"❌ Proposal {PROPOSAL_ID} not found!")
        return None

    return proposal


def analyze_excel(excel_path):
    """Analyze the exported Excel file."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    print("=" * 80)
    print("EXCEL FILE ANALYSIS")
    print("=" * 80)

    print(f"\n📊 Sheets: {len(wb.sheetnames)}")
    for idx, name in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {name}")

    # CE Summary
    ws_ce = wb["CE Summary"]
    print(f"\n💰 CE Summary (Base Period - Column C):")
    print(f"  Direct Labor (C10): {ws_ce['C10'].value}")
    print(f"  Fringe (C11): {ws_ce['C11'].value}")
    print(f"  Overhead (C12): {ws_ce['C12'].value}")
    print(f"  G&A (C13): {ws_ce['C13'].value}")
    print(f"  Subcontractors (C14): {ws_ce['C14'].value}")
    print(f"  S&MH (C15): {ws_ce['C15'].value}")

    # Prime Labor Detail
    ws_labor = wb["Prime Labor Detail"]
    print(f"\n👥 Prime Labor Detail:")
    print(f"  First Position: {ws_labor['B11'].value}")
    print(f"  Location: {ws_labor['D11'].value}")

    # Check for FBLR breakdown
    print(f"\n  FBLR Breakdown Rows:")
    for row in range(20, min(40, ws_labor.max_row + 1)):
        label = ws_labor.cell(row, 2).value
        if label and any(kw in str(label) for kw in ['Total Direct', 'Fringe', 'Overhead', 'G&A', 'Fee']):
            value = ws_labor.cell(row, 6).value  # First amount column
            print(f"    {label}: {value}")

    # LOE Company column
    ws_loe = wb["LOE"]
    print(f"\n🏢 LOE Sheet - Company Column:")
    print(f"  Header C10: {ws_loe['C10'].value}")
    print(f"  First Position Company (C11): {ws_loe['C11'].value}")

    wb.close()


def analyze_proposal(proposal):
    """Analyze proposal data from MongoDB."""
    print("\n" + "=" * 80)
    print("MONGODB PROPOSAL DATA")
    print("=" * 80)

    spreadsheet_data = proposal.get('spreadsheet_data', {})

    print(f"\n📄 Proposal Info:")
    print(f"  Name: {proposal.get('name', 'N/A')}")
    print(f"  Prime Contractor: {spreadsheet_data.get('prime_contractor_name', 'N/A')}")
    print(f"  Solicitation: {spreadsheet_data.get('solicitation_number', 'N/A')}")

    # Rates
    rates = spreadsheet_data.get('rates', {})
    print(f"\n💹 Indirect Rates:")
    print(f"  Fringe: {rates.get('fringe', 'N/A')}")
    print(f"  OH Onsite: {rates.get('oh_onsite', rates.get('oh', 'N/A'))}")
    print(f"  OH Offsite: {rates.get('oh_offsite', rates.get('oh', 'N/A'))}")
    print(f"  G&A: {rates.get('ga', 'N/A')}")
    print(f"  Fee: {rates.get('fee', 'N/A')}")

    # Positions
    positions = spreadsheet_data.get('positions', [])
    print(f"\n👥 Positions: {len(positions)} total")

    # Count prime vs subcontractor
    prime_count = sum(1 for p in positions if not p.get('assigned_subcontractor_id'))
    sub_count = sum(1 for p in positions if p.get('assigned_subcontractor_id'))
    print(f"  Prime positions: {prime_count}")
    print(f"  Subcontractor positions: {sub_count}")

    if positions:
        # First prime position
        prime_pos = next((p for p in positions if not p.get('assigned_subcontractor_id')), None)
        if prime_pos:
            print(f"\n  First Prime Position:")
            print(f"    Labor Category: {prime_pos.get('labor_category', 'N/A')}")
            print(f"    Location: {prime_pos.get('location', 'N/A')}")
            print(f"    Year 1 Hours: {prime_pos.get('hours_per_year', {}).get('1', 0)}")
            print(f"    Year 1 Rate: ${prime_pos.get('year_1_rate', 0):,.2f}")
            print(f"    Year 1 Amount: ${prime_pos.get('year_1_amount', 0):,.2f}")

    # Subcontractors
    subcontractors = spreadsheet_data.get('subcontractors', [])
    print(f"\n🏢 Subcontractors: {len(subcontractors)}")
    for idx, sub in enumerate(subcontractors, 1):
        sub_positions = [p for p in positions if p.get('assigned_subcontractor_id') == sub.get('id')]
        print(f"  {idx}. {sub.get('name', 'N/A')} - {len(sub_positions)} positions")

    # Aggregates
    aggregates = spreadsheet_data.get('aggregates', {})
    if aggregates:
        year_1 = aggregates.get('year_1', {})
        print(f"\n💰 Year 1 Aggregates:")
        print(f"  Direct Labor: ${year_1.get('direct_labor', 0):,.2f}")
        print(f"  Fringe: ${year_1.get('fringe', 0):,.2f}")
        print(f"  Overhead: ${year_1.get('overhead', 0):,.2f}")
        print(f"  G&A: ${year_1.get('ga', 0):,.2f}")
        print(f"  Subcontractors: ${year_1.get('subcontractors', 0):,.2f}")
        print(f"  Total: ${year_1.get('total', 0):,.2f}")


def compare_values(proposal, excel_path):
    """Compare MongoDB values with Excel values."""
    print("\n" + "=" * 80)
    print("COMPARISON - MongoDB vs Excel")
    print("=" * 80)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws_ce = wb["CE Summary"]

    spreadsheet_data = proposal.get('spreadsheet_data', {})
    aggregates = spreadsheet_data.get('aggregates', {})
    year_1 = aggregates.get('year_1', {}) if aggregates else {}

    # Compare Direct Labor
    mongo_dl = year_1.get('direct_labor', 0)
    excel_dl = ws_ce['C10'].value

    print(f"\n🔍 Direct Labor (Year 1):")
    print(f"  MongoDB: ${mongo_dl:,.2f}")
    print(f"  Excel C10: {excel_dl}")

    if isinstance(excel_dl, (int, float)):
        diff = abs(mongo_dl - excel_dl)
        if diff < 0.01:
            print(f"  ✅ MATCH (diff: ${diff:.4f})")
        else:
            print(f"  ❌ MISMATCH (diff: ${diff:,.2f})")
    else:
        print(f"  ⚠️  Excel value is not a number (probably formula not calculated)")

    # Compare Fringe
    mongo_fringe = year_1.get('fringe', 0)
    excel_fringe = ws_ce['C11'].value

    print(f"\n🔍 Fringe Benefits (Year 1):")
    print(f"  MongoDB: ${mongo_fringe:,.2f}")
    print(f"  Excel C11: {excel_fringe}")

    if isinstance(excel_fringe, (int, float)):
        diff = abs(mongo_fringe - excel_fringe)
        if diff < 0.01:
            print(f"  ✅ MATCH (diff: ${diff:.4f})")
        else:
            print(f"  ❌ MISMATCH (diff: ${diff:,.2f})")

    wb.close()


def main():
    """Main verification function."""
    print("=" * 80)
    print("PROPOSAL VERIFICATION SCRIPT")
    print("=" * 80)
    print(f"\nProposal ID: {PROPOSAL_ID}")
    print(f"Excel File: {EXCEL_PATH}")
    print(f"MongoDB URL: {settings.MONGODB_URL}")
    print(f"Database: {settings.MONGODB_DATABASE}")

    # Fetch proposal
    proposal = fetch_proposal_from_mongo()
    if not proposal:
        return

    # Analyze both
    analyze_proposal(proposal)
    analyze_excel(EXCEL_PATH)

    # Compare
    compare_values(proposal, EXCEL_PATH)

    print("\n" + "=" * 80)
    print("VERIFICATION COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    main()
